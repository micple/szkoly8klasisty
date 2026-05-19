from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import quote, urlencode

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "wybor_szkol_warszawa.xlsx"
SITE_DIR = ROOT / "site"


def commons_file(name: str) -> str:
    return f"https://commons.wikimedia.org/wiki/Special:FilePath/{quote(name)}?width=1200"


def gallery_url(item: str) -> str:
    if item.startswith("http://") or item.startswith("https://"):
        return item
    return commons_file(item)


def image_search_url(name: str) -> str:
    return "https://www.google.com/search?" + urlencode({"tbm": "isch", "q": f"{name} Warszawa budynek zdjęcia"})


def opinion_search_url(name: str) -> str:
    return "https://www.google.com/search?" + urlencode({"q": f"{name} Warszawa opinie forum reddit gazeta"})


ENRICHMENT = {
    "VIII LO im. Władysława IV": {
        "official": "https://wladyslaw.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:VIII_LO_im._W%C5%82adys%C5%82awa_IV",
        "summary": "Historyczne, selektywne liceum na Pradze-Północ; bardzo dobre połączenie z Kobyłką i mocne profile ścisło-przyrodnicze.",
        "opinion": "Warto sprawdzić tempo pracy i atmosferę w klasach biol-chem/ścisłych; wysoki próg sugeruje ambitne środowisko.",
        "gallery": [
            "VIII Liceum Ogólnokształcące im. Władysława IV w Warszawie 2021.jpg",
            "VIII Liceum Ogólnokształcące im. Władysława IV w Warszawie 2017.jpg",
            "VIII LO im. Władysława IV w Warszawie 2015.JPG",
        ],
    },
    "XXVII LO im. Tadeusza Czackiego": {
        "official": "https://czacki.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XXVII_LO_im._Tadeusza_Czackiego",
        "summary": "Jedna z najmocniejszych śródmiejskich szkół; bardzo wysoka selektywność i profile zgodne z medycyną oraz politechniką.",
        "opinion": "Do sprawdzenia poziom presji i obciążenia pracą; bardzo wysokie progi oznaczają mocne, konkurencyjne środowisko.",
        "gallery": [
            "XXVII Liceum Ogólnokształcące im. Tadeusza Czackiego w Warszawie 2017.jpg",
            "Czacki1.jpg",
            "Czacki2.jpg",
        ],
    },
    "IX LO im. Klementyny Hoffmanowej": {
        "official": "https://www.dziewiatka.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:IX_LO_im._Klementyny_Hoffmanowej",
        "summary": "Bardzo mocne liceum śródmiejskie, szczególnie atrakcyjne dla profili biol-chem-mat i chem-fiz-mat.",
        "opinion": "Opinie internetowe warto czytać ostrożnie: powtarza się motyw wysokich wymagań i ambitnych uczniów.",
        "gallery": [
            "IX Liceum Ogólnokształcące im. Klementyny Hoffmanowej 2025.jpg",
            "IX Liceum Ogólnokształcące im. Klementyny Hoffmanowej w Warszawie 2018.jpg",
            "IX L.O. im. Hoffmanowej 03.jpg",
        ],
    },
    "II LO im. Stefana Batorego": {
        "official": "https://batory.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:II_LO_im._Stefana_Batorego_w_Warszawie",
        "summary": "Elitarne liceum w centrum; szczególnie mocne jako ambitny wybór pod medycynę i kierunki ścisłe.",
        "opinion": "Na forach pojawiają się dyskusje o wysokim poziomie i dużej ambicji uczniów; warto pytać obecnych uczniów o realne obciążenie.",
        "gallery": [
            "II Liceum Ogólnokształcące z Oddziałami Dwujęzycznymi im. Stefana Batorego w Warszawie 2025.jpg",
            "II Liceum Ogólnokształcące z Oddziałami Dwujęzycznymi im. Stefana Batorego w Warszawie 2025a.jpg",
            "Batory patio.jpg",
        ],
    },
    "XVIII LO im. Jana Zamoyskiego": {
        "official": "https://zamoyski.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XVIII_LO_im._Jana_Zamoyskiego_w_Warszawie",
        "summary": "Śródmiejskie liceum z mocnymi profilami ścisłymi; dobry kompromis między selektywnością a dojazdem.",
        "opinion": "Warto sprawdzić opinie o konkretnych nauczycielach rozszerzeń, zwłaszcza biologii, chemii i matematyki.",
        "gallery": [
            "XVIII Liceum Ogólnokształcące im. Jana Zamoyskiego w Warszawie 2017.jpg",
            "18 LO in Warsaw.jpg",
            "High school XVIII, Warsaw city.jpg",
        ],
    },
    "V LO im. Księcia Józefa Poniatowskiego": {
        "official": "https://vlo.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:V_LO_im._Ks._J%C3%B3zefa_Poniatowskiego_w_Warszawie",
        "summary": "Dobre śródmiejskie liceum pod profile medyczne i pomostowe chem-fiz-mat.",
        "opinion": "Sprawdzić balans między wymaganiami a wsparciem; wysokie progi sugerują klasę ambitną, ale nie ekstremalnie niedostępną.",
        "gallery": [
            "V Liceum Ogólnokształcące im. Księcia Józefa Poniatowskiego w Warszawie 2023.jpg",
            "V Liceum Ogólnokształcące im. Ks. Józefa Poniatowskiego w Warszawie 2016.jpg",
            "V Liceum Ogólnokształcące im. Księcia Józefa Poniatowskiego, ul. Nowolipie 8, Warszawa.jpg",
        ],
    },
    "XXXVII LO im. Jarosława Dąbrowskiego": {
        "official": "https://dabrowski37lo.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/File:XXXVII_Liceum_Og%C3%B3lnokszta%C5%82c%C4%85ce_im._Jaros%C5%82awa_D%C4%85browskiego_-_panoramio.jpg",
        "summary": "Bardzo sensowna śródmiejska opcja dla profili fiz-inf-mat i biol-chem-mat, z dojazdem metrem.",
        "opinion": "Źródła szkolne podkreślają aktywność uczniów i projekty; fora warto wykorzystać do oceny atmosfery i tempa pracy.",
        "gallery": [
            "XXXVII Liceum Ogólnokształcące im. Jarosława Dąbrowskiego - panoramio.jpg",
            "https://dabrowski37lo.edu.pl/wp-content/uploads/2018/03/3.png",
            "https://dabrowski37lo.edu.pl/wp-content/uploads/2026/04/02_szkola_sukcesu.png",
        ],
    },
    "XI LO im. Mikołaja Reja": {
        "official": "https://www.rej.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XI_LO_im._Miko%C5%82aja_Reja_w_Warszawie",
        "summary": "Śródmiejskie liceum z profilem matematyczno-fizycznym oraz biol-chem; dobre jako ambitny-realny wybór.",
        "opinion": "Strona szkoły akcentuje projekty, aktywność społeczną i profile klas; warto dopytać o poziom konkretnych rozszerzeń.",
        "gallery": [
            "XI Liceum Ogólnokształcące im. Mikołaja Reja w Warszawie 2025.jpg",
            "XI Liceum Ogólnokształcące im. Mikołaja Reja w Warszawie.jpg",
        ],
    },
    "XIII LO im. płk. Leopolda Lisa-Kuli": {
        "official": "https://lo13targowek.eduwarszawa.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XIII_LO_im._p%C5%82k._Leopolda_Lisa-Kuli",
        "summary": "Bardzo praktyczny wybór blisko Kobyłki; dobre klasy mat-inf/fiz i medyczne jako realny środek listy.",
        "opinion": "Warto sprawdzić opinie lokalne o atmosferze i dojazdach z Marek/Ząbek/Kobyłki; punktowo to rozsądny bufor.",
        "gallery": [
            "XIII LO Lisa-Kuli.jpg",
            "XIII LO w Warszawie.jpg",
            "Liceum Lisa-Kuli .jpg",
        ],
    },
    "XXXV LO im. Bolesława Prusa": {
        "official": "https://prus.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XXXV_LO_im._Boles%C5%82awa_Prusa",
        "summary": "Mocna praska szkoła z profilem fiz-ang-mat; dobry wybór na styku ambitne-realne.",
        "opinion": "Do sprawdzenia opinie o profilu ścisłym i językach; progi wskazują na wysoką, ale nie ekstremalną selektywność.",
        "gallery": [
            "XXXV Liceum Ogólnokształcące im. Bolesława Prusa w Warszawie 2016.JPG",
            "XXXV Liceum Ogólnokształcące im. Bolesława Prusa w Warszawie kwiecień 2019.jpg",
            "Syrenka XXXV Liceum Ogólnokształcące z Oddziałami Dwujęzycznymi im. Bolesława Prusa w Warszawie ul. Zwycięzców 7-9.jpg",
        ],
    },
    "LXXXIII LO im. Emiliana Konopczyńskiego": {
        "official": "https://konopczynski.com/",
        "summary": "Śródmiejska opcja bardziej realna punktowo; dobra jako zabezpieczenie w centrum dla mat-fiz lub biol-chem-mat.",
        "opinion": "Warto porównać opinie o poziomie rozszerzeń z bardziej selektywnymi liceami; może być dobrym spokojniejszym wyborem.",
        "gallery": [
            "https://konopczynski.com/wp-content/uploads/2021/09/zdj1-1024x576.png",
            "https://konopczynski.com/wp-content/uploads/2021/09/konop-1024x427.png",
            "https://konopczynski.com/wp-content/uploads/2026/04/677204382_1515198497283326_6996017545879373484_n-300x169.jpg",
        ],
    },
    "CLVII LO im. Marii Skłodowskiej-Curie": {
        "official": "http://www.lo157.waw.pl/",
        "summary": "Śródmiejskie liceum z bardzo sensownymi profilami fiz-inf-mat i chem/biol-mat; dobre połączenie z centrum.",
        "opinion": "Warto dopytać o konkretne klasy i nauczycieli rozszerzeń; punktowo to realna, ale ambitna opcja.",
        "gallery": [
            "https://lo157.waw.pl/wp-content/uploads/2026/04/1000011875-2048x1536.jpg",
            "https://lo157.waw.pl/wp-content/uploads/2025/04/tarcze-nowe-2025-zlote-liceum-edited-scaled.jpg",
            "https://lo157.waw.pl/wp-content/uploads/2026/02/zlota-szkola-licea-1024x1024.png",
        ],
    },
    "Technikum Mechatroniczne nr 1": {
        "official": "https://tm1.edu.pl/",
        "summary": "Najbardziej naturalny wybór techniczny pod mechatronikę, robotykę, automatykę i politechnikę.",
        "opinion": "Kluczowe pytanie to dojazd z Kobyłki oraz to, czy syn chce pięcioletniej ścieżki zawodowej zamiast licealnej matury.",
        "gallery": [
            "https://tm1.edu.pl/wp-content/uploads/2024/09/bufet-wisniowa56-scaled.jpg",
            "https://tm1.edu.pl/wp-content/uploads/2025/07/boisko.jpg",
            "https://tm1.edu.pl/wp-content/uploads/2025/07/strefa-kandydata.jpg",
        ],
    },
    "VI LO im. Tadeusza Reytana": {
        "official": "https://reytan.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:VI_LO_im._Tadeusza_Reytana",
        "summary": "Mocne liceum mokotowskie z profilami ścisłymi; merytorycznie atrakcyjne, ale dojazdowo cięższe.",
        "opinion": "Warto traktować jako ambitny wybór, jeśli profil i renoma przeważą nad codziennym dojazdem.",
        "gallery": [
            "VI Liceum Ogólnokształcące im. Tadeusza Reytana w Warszawie 2020.jpg",
            "VI LO im. Tadeusza Reytana w Warszawie.JPG",
            "VI Liceum Ogólnokształcące im. Tadeusza Reytana w Warszawie 2025.jpg",
        ],
    },
    "Technikum Kinematograficzno-Komputerowe im. K. Kieślowskiego": {
        "official": "https://technikumpolna.pl/",
        "summary": "Techniczna opcja programistyczna w Śródmieściu; dobra pod IT, mniej uniwersalna pod medycynę klasyczną.",
        "opinion": "Sprawdzić realną jakość przedmiotów zawodowych i praktyk; progi są bardziej dostępne niż Wiśniowa.",
        "gallery": [
            "https://technikumpolna.pl/wp-content/uploads/2025/02/20200515_134139-2.jpg",
            "https://technikumpolna.pl/wp-content/uploads/2025/02/maxresdefault.jpg",
            "https://technikumpolna.pl/wp-content/uploads/2025/02/20200515_134139-2-1.jpg",
        ],
    },
    "XIX LO im. Powstańców Warszawy": {
        "official": "https://lo19.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:XIX_LO_im._Powsta%C5%84c%C3%B3w_Warszawy",
        "summary": "Praska opcja ścisła, punktowo blisko granicy 168; sensowna, jeśli dojazd jest akceptowalny.",
        "opinion": "Warto porównać z Prusem i Lisa-Kuli pod kątem atmosfery i rozszerzeń.",
        "gallery": [
            "XIX Liceum Ogólnokształcące im. Powstańców Warszawy w Warszawie 2019.jpg",
            "https://lo19.pl/rekrutacja.jpg",
            "https://lo19.pl/img_8474.jpeg",
        ],
    },
    "CV LO im. Zbigniewa Herberta": {
        "official": "https://cvlo.waw.pl/",
        "commons": "https://commons.wikimedia.org/wiki/Category:CV_LO_im._Zbigniewa_Herberta",
        "summary": "Białołęcka opcja zapasowa dla profili fiz-ang-mat i biol-chem; bardziej bezpieczna punktowo.",
        "opinion": "Dobra do zabezpieczenia listy, ale warto sprawdzić dojazd i poziom rozszerzeń względem szkół z centrum.",
        "gallery": [
            "2025 Warszawa CV Liceum Ogólnokształcące im. Zbigniewa Herberta, 1.jpg",
            "https://cvlo.waw.pl/images/slider/sala-lekcyjna-145-listopad.jpg",
            "https://cvlo.waw.pl/images/slider/boisko.jpg",
        ],
    },
    "Technikum Elektroniczne nr 1": {
        "official": "https://elektronik.edu.pl/",
        "summary": "Techniczny zapas dla mechatroniki/elektroniki; sensowny pod politechnikę i aparaturę medyczną.",
        "opinion": "Sprawdzić praktyki zawodowe, bazę techniczną i dojazd na Wolę.",
        "gallery": [
            "https://elektronik.edu.pl/images/grafika-menu/egzamin_zawodowy.jpg",
            "https://elektronik.edu.pl/images/grafika-menu/egzamin_maturalny.jpg",
            "https://elektronik.edu.pl/images/tm/technik_cyber_small.png",
        ],
    },
    "XII LO im. Henryka Sienkiewicza": {
        "official": "https://12lo-warszawa.edupage.org/",
        "summary": "Profil geogr-ang-mat jest mniej celowany w medycynę, ale może pasować do kierunków techniczno-ekonomicznych.",
        "opinion": "Raczej wybór pomocniczy niż główny pod zadane cele.",
        "gallery": [
            "Warsaw Ghetto boundary marker 53 Sienna Street.JPG",
            "https://12lo-warszawa.edupage.org/global/pics/skins/slide1900/thumbs/fill320x320trslide30.jpg",
            "https://12lo-warszawa.edupage.org/global/pics/skins/slide/thumbs/fill320x320trslide101.jpg",
        ],
    },
    "Technikum Łączności": {
        "official": "https://zsl.waw.pl/typy-szkol/technikum/",
        "summary": "Bezpieczniejszy techniczny wybór pod informatykę/telekomunikację; słabszy prestiżowo od pierwszych techników.",
        "opinion": "Warto zweryfikować praktyki i realny poziom programowania.",
        "gallery": [
            "https://zsl.waw.pl/wp-content/uploads/2021/01/logo-1.png",
            "https://zsl.waw.pl/images/stories/zs37/cisco_logo_large.jpg",
            "https://zsl.waw.pl/wp-content/uploads/2026/05/HackCarpathia_2026_04.jpg",
        ],
    },
    "II LO im. Stefana Batorego - IB": {
        "official": "https://batory.edu.pl/",
        "summary": "Publiczna ścieżka IB/pre-IB w jednym z najmocniejszych liceów w Warszawie; bardzo dobra przy planie studiów zagranicznych.",
        "opinion": "Do sprawdzenia: koszt programu IB, obciążenie pracą, język wykładowy i dostępność przedmiotów HL pod STEM/medycynę.",
        "gallery": [
            "II Liceum Ogólnokształcące z Oddziałami Dwujęzycznymi im. Stefana Batorego w Warszawie 2025.jpg",
            "II Liceum Ogólnokształcące z Oddziałami Dwujęzycznymi im. Stefana Batorego w Warszawie 2025a.jpg",
            "Batory patio.jpg",
        ],
    },
    "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika": {
        "official": "https://kopernik.edu.pl/",
        "summary": "Bardzo mocne publiczne liceum dwujęzyczne na Woli, szczególnie sensowne dla ścieżki politechnicznej i medycznej.",
        "opinion": "To wybór ambitny: progi są bardzo wysokie, a oddziały dwujęzyczne zwykle wymagają sprawdzianu kompetencji językowych. Warto dać wysoko, jeśli syn akceptuje ryzyko i intensywny angielski.",
        "gallery": [
            "XXXIII Liceum Ogólnokształcące Dwujęzyczne im. Mikołaja Kopernika w Warszawie 2019.jpg",
        ],
    },
    "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika - IB/MYP": {
        "official": "https://kopernik.edu.pl/program-dyplomowy-matury-miedzynarodowej/",
        "summary": "Bardzo mocna publiczna szkoła IB/MYP/DP na Woli; jedna z najbardziej sensownych opcji IB pod ambitne studia.",
        "opinion": "Największe ryzyko to bardzo wysoki próg i wymagający angielski; warto sprawdzić wybór HL Math, Physics, Chemistry i Biology.",
        "gallery": [
            "XXXIII Liceum Ogólnokształcące Dwujęzyczne im. Mikołaja Kopernika w Warszawie 2019.jpg",
        ],
    },
    "XXXV LO im. Bolesława Prusa - IB": {
        "official": "https://prus.edu.pl/ib-dp/program-ib-dp/",
        "summary": "Warszawska ścieżka IB w znanej publicznej szkole po prawej stronie Wisły; logistycznie ciekawa alternatywa dla centrum.",
        "opinion": "Sprawdzić, czy aktualna oferta przedmiotów HL pasuje do medycyny albo politechniki; IB wymaga innego stylu pracy niż polska matura.",
        "gallery": [
            "XXXV Liceum Ogólnokształcące im. Bolesława Prusa w Warszawie 2016.JPG",
            "XXXV Liceum Ogólnokształcące im. Bolesława Prusa w Warszawie kwiecień 2019.jpg",
        ],
    },
    "2 SLO z Oddz. Międzynarodowymi im. P. Jasienicy STO": {
        "official": "https://ib.2slo.pl/",
        "summary": "Bardzo mocna społeczna szkoła IB w centrum; świetna pod studia zagraniczne i ambitne środowisko.",
        "opinion": "Wymaga osobnej decyzji finansowej i sprawdzenia kultury pracy; warto porównać z publicznym Kopernikiem i Batorym.",
        "gallery": [
            "https://ib.2slo.pl/wp-content/uploads/2025/03/heroib2025-scaled-e1741008077904-1920x930.jpg",
            "https://ib.2slo.pl/wp-content/uploads/2026/05/IMG_1408-scaled-e1778494081150.jpg",
            "https://ib.2slo.pl/wp-content/uploads/2026/04/TEDx2slo-photo-all-togetherAutorNieznany-scaled.jpg",
        ],
    },
    "Prywatne LO im. Zofii i Jędrzeja Moraczewskich Monnet International School": {
        "official": "https://www.maturamiedzynarodowa.pl/liceum/",
        "summary": "Prywatna szkoła IB wyspecjalizowana w ścieżce międzynarodowej; dobra, jeśli priorytetem są studia zagraniczne.",
        "opinion": "Do osobnej weryfikacji: koszt, dojazd z Kobyłki i realna oferta HL dla nauk ścisłych/przyrodniczych.",
        "gallery": [
            "https://www.maturamiedzynarodowa.pl/liceum/wp-content/uploads/sites/2/2024/08/0_0-6-690x310.png",
            "https://www.maturamiedzynarodowa.pl/liceum/wp-content/uploads/sites/2/2024/08/0_2-690x310.png",
            "https://www.maturamiedzynarodowa.pl/liceum/wp-content/uploads/sites/2/2026/04/Multimedia-5-480x320.jpeg",
        ],
    },
    "International American School of Warsaw": {
        "official": "https://ias.edu.pl/",
        "summary": "Anglojęzyczna szkoła międzynarodowa z IB DP; bardziej ścieżka globalna niż klasyczna warszawska rekrutacja LO.",
        "opinion": "Sensowna tylko, jeśli rodzina akceptuje koszt, model prywatny i silne przesunięcie w stronę edukacji międzynarodowej.",
        "gallery": [
            "https://ias.edu.pl/wp-content/uploads-new/2026/03/ias-og-image-1.png",
            "https://ias.edu.pl/wp-content/uploads-new/2023/02/JWA_20180621_IAS_00193-scaled.jpg",
            "https://ias.edu.pl/wp-content/uploads-new/2023/02/36339964_1746670632077415_615550417895424_o.jpg",
        ],
    },
    "Warsaw Montessori High School": {
        "official": "https://highschool.wmf.edu.pl/ib-diploma-programme/",
        "summary": "Kameralna prywatna szkoła z Pre-IB i IB DP w centrum; potencjalnie dobra przy potrzebie bardziej indywidualnego podejścia.",
        "opinion": "Warto dokładnie sprawdzić poziom i dostępność przedmiotów HL pod politechnikę lub medycynę.",
        "gallery": [
            "https://highschool.wmf.edu.pl/wp-content/uploads/sites/8/2025/01/Warsaw-Montessori-High-School-–-kopia-1080-x-1080-px.jpg",
            "https://highschool.wmf.edu.pl/wp-content/uploads/sites/8/2024/09/ib-world-school-logo-2-colour.png",
            "https://highschool.wmf.edu.pl/wp-content/uploads/sites/8/2024/10/dp-model-en-1-1024x1024.png",
        ],
    },
    "Thames British School Warsaw": {
        "official": "https://thamesbritishschool.pl/learning/diploma-program/",
        "summary": "Prywatna brytyjska szkoła międzynarodowa z IB DP i szeroką ofertą przedmiotów, w tym Math, Physics, Chemistry, Biology i Computer Science.",
        "opinion": "Merytorycznie ciekawa dla STEM/med, ale dojazd z Kobyłki i koszt są dużymi czynnikami decyzyjnymi.",
        "gallery": [
            "https://thamesbritishschool.pl/wp/../app/uploads/2025/02/Projekt-bez-nazwy-9-540x0-c-default.png",
            "https://thamesbritishschool.pl/wp/../app/uploads/2025/01/320-x-433-our-2-22-320x0-c-default.png",
            "https://thamesbritishschool.pl/app/themes/osom-theme/src/img//decoration-learning.png",
        ],
    },
    "The British School Warsaw": {
        "official": "https://www.nordangliaeducation.com/our-schools/warsaw/ib",
        "summary": "Międzynarodowa szkoła z długo prowadzonym IB DP; najmocniej pasuje do scenariusza studiów zagranicznych.",
        "opinion": "To raczej strategiczny wybór międzynarodowy niż zamiennik dobrego profilu mat-fiz/biol-chem w publicznym LO.",
        "gallery": [
            "https://www.nordangliaeducation.com/tbsw-warsaw/-/media/aprimo/1dd029/bswwarsawsept-2021146tbswcurriculadetailiboptterheader1317x4392.jpg?h=439&iar=0&w=1317&rev=22353dcfa79843dda8975b6ab9775497&hash=D54F2D81F7846F200F269B32AF2471D1",
            "https://www.nordangliaeducation.com/tbsw-warsaw/-/media/warsaw/academic-excellence/curricula/tbsw-ib-2.png?h=900&iar=0&w=720&rev=c0d3d4b4724d4fa2986b88449601c130&hash=6C4A46E0A34206A3AFAFDADB22FB3A99",
            "https://www.nordangliaeducation.com/tbsw-warsaw/-/media/aprimo/499ae8/bsw_warsaw_sept-2021_102_tbsw_ib_academicexcell_vistimeline_374x249_2.jpg?h=252&iar=0&w=372&rev=e39a4c81a50847ca97914ac5eaee67c5&hash=EEE8612C8262243F56AB8E40BBD838D5",
        ],
    },
    "Międzynarodowe LO TE Vizja Warszawa Centrum": {
        "official": "https://tevizja.pl/warszawa-centrum/program-ib/",
        "summary": "Prywatna ścieżka IB w centrum Warszawy; logistycznie sensowna przy dojeździe metrem.",
        "opinion": "Sprawdzić jakość przedmiotów HL, stabilność programu i koszty przed wpisaniem wysoko na listę.",
        "gallery": [
            "https://tevizja.pl/wp-content/uploads/2023/11/LO_05.jpg",
            "https://tevizja.pl/wp-content/uploads/2023/11/ib-diploma.png",
            "https://tevizja.pl/wp-content/themes/tevizja/images/logoLO_new_footer.png",
        ],
    },
    "Prywatne LO Sióstr Nazaretanek z Oddziałami Międzynarodowymi": {
        "official": "https://nazaretanki.edu.pl/program-matury-miedzynarodowej-ib-dp/",
        "summary": "Prywatna ścieżka pre-IB oraz IB DP; opcja międzynarodowa z własną rekrutacją.",
        "opinion": "Merytorycznie warto rozważyć, ale dojazd z Kobyłki na Wilanów jest słaby.",
        "gallery": [
            "https://nazaretanki.edu.pl/wp-content/uploads/2025/08/250627_Nazaret_zak_rok_0524-Poprawione-Szum-scaled.jpg",
            "https://nazaretanki.edu.pl/wp-content/uploads/2023/03/ib_1x1_1080_1_@2x.webp",
            "https://nazaretanki.edu.pl/wp-content/uploads/2023/02/Nazaret_Logo.svg",
        ],
    },
    "CLXI LO im. Władysława Bartoszewskiego": {
        "official": "https://www.lo161.waw.pl/",
        "summary": "Buforowa szkoła na Woli z progami wyraźnie poniżej 150 pkt; dobra lokalizacja i profile mat/geogr oraz biol-chem-ang.",
        "opinion": "Traktować jako zabezpieczenie listy, nie wybór aspiracyjny; warto sprawdzić realny poziom rozszerzeń.",
        "gallery": [],
    },
    "XLVI LO im. Stefana Czarnieckiego": {
        "official": "https://lo46targowek.eduwarszawa.pl/",
        "summary": "Targówek, dobry dojazd z Kobyłki i progi poniżej 150 pkt; sensowny dolny bufor z profilami przyrodniczymi i matematycznymi.",
        "opinion": "Warto sprawdzić atmosferę i poziom konkretnych nauczycieli rozszerzeń, bo to wybór zabezpieczający.",
        "gallery": [],
    },
    "L LO im. Ruy Barbosy": {
        "official": "https://lo50.edu.pl/",
        "commons": "https://commons.wikimedia.org/wiki/File:L_Liceum_Og%C3%B3lnokszta%C5%82c%C4%85ce_z_Oddzia%C5%82ami_Integracyjnymi_im._Ruy_Barbosy_w_Warszawie_2020.jpg",
        "summary": "Praga-Północ, bardzo dobry bufor logistyczny; szczególnie ciekawy jest profil fiz-inf-mat poniżej 150 pkt.",
        "opinion": "Najlepszy z dodanych buforów pod politechnikę, jeśli priorytetem jest pewność dostania i sensowny dojazd.",
        "gallery": [
            "L Liceum Ogólnokształcące z Oddziałami Integracyjnymi im. Ruy Barbosy w Warszawie 2020.jpg",
            "https://lo50.edu.pl/templates/lo50/images/logo.png",
        ],
    },
    "XLV LO im. Romualda Traugutta": {
        "official": "http://www.traugutt.edu.pl/",
        "summary": "Wolski bufor z profilem inf-ang-mat poniżej 150 pkt; dobry jako zabezpieczenie informatyczno-matematyczne.",
        "opinion": "Sprawdzić poziom informatyki i matematyki, bo próg jest bezpieczny, ale ambicja profilu zależy od szkoły i nauczycieli.",
        "gallery": [
            "http://cloud-4.edupage.org/cloud?z%3AHXE%2FBpga8jzBrSxr%2BLOEMB29S7fFg2fHDS%2B39z6kSVxpdN6dZ3x2NlXFU61bM9jC8QVUQVM4g7WcwAqDsyh73w%3D%3D",
            "http://cloud-d.edupage.org/cloud?z%3AEfWQLRJjGkTAttlV0qB%2F5msm8wUZT1t9XOZ1U2TQdPg2s5EHhg%2FBQuq4r7ScJFXU2YgEwefqviTGsXii3WsVJYWXl2f5HpyLZP7HDrClECE%3D",
        ],
    },
    "LXXXVI LO im. Batalionu Zośka": {
        "official": "https://zoska.waw.pl/",
        "summary": "Wola, profil fiz-ang-mat poniżej 150 pkt; sensowna bezpieczna opcja pod matematykę i fizykę.",
        "opinion": "Dobry bufor, ale do sprawdzenia, czy poziom rozszerzeń wystarczy pod ambitniejszą politechnikę.",
        "gallery": [
            "https://zoska.waw.pl/wp-content/uploads/2023/04/strona-szkoly.png",
            "https://zoska.waw.pl/wp-content/uploads/2021/06/Zrzut-ekranu-2021-06-07-142056.png",
        ],
    },
    "Technikum Geologiczno-Geodezyjno-Drogowe": {
        "official": "http://www.zs14.pl/",
        "summary": "Techniczny bufor na Pradze-Północ z kierunkiem programista i bardzo niskim progiem; dobra lokalizacja awaryjna.",
        "opinion": "Opcja zapasowa, jeśli celem jest utrzymanie ścieżki technicznej; trzeba zweryfikować jakość programowania i praktyk.",
        "gallery": [
            "http://www.zs14.pl/images/1._PLIKI_SZABLONU_STRONY/gowne.jpg",
            "http://www.zs14.pl/images/technik_programista.jpg",
            "http://www.zs14.pl/images/technik_geodeta.jpg",
        ],
    },
    "LXXVI LO im. Marszałka Józefa Piłsudskiego": {
        "official": "https://pilsudski.edu.pl/",
        "summary": "Praga-Północ, bardzo bezpieczny próg i rozsądny dojazd; profil geogr-ang-mat bardziej techniczno-ekonomiczny niż medyczny.",
        "opinion": "Traktować jako ostatni bezpieczny bufor w dobrej lokalizacji, nie jako szkołę główną pod medycynę.",
        "gallery": [
            "https://pilsudski.edu.pl/global/pics/skins/slide/thumbs/fill320x320trslide102.jpg",
            "https://pilsudski.edu.pl/global/pics/skins/slide/thumbs/fill320x320trslide103.jpg",
            "https://pilsudski.edu.pl/global/pics/skins/slide/thumbs/fill320x320trslide101.jpg",
        ],
    },
}


def rows_from_workbook() -> list[dict]:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Rekomendacja"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def build_school_records(rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        school = row["Szkoła"]
        grouped.setdefault(
            school,
            {
                "school": school,
                "district": row["Dzielnica"],
                "bestScore": row["Atrakcyjność 0-100"],
                "bestRisk": row["Szansa przy 168 pkt"],
                "classes": [],
            },
        )
        grouped[school]["bestScore"] = max(grouped[school]["bestScore"], row["Atrakcyjność 0-100"])
        grouped[school]["classes"].append(row)

    records = []
    for school, record in grouped.items():
        record["classes"].sort(key=lambda item: item["Atrakcyjność 0-100"], reverse=True)
        top = record["classes"][0]
        extra = ENRICHMENT.get(school, {})
        gallery_files = extra.get("gallery", [])
        record.update(
            {
                "official": extra.get("official", top.get("Źródło IB", "") or ""),
                "commons": extra.get("commons", ""),
                "summary": extra.get("summary", "Opis szkoły wymaga ręcznej weryfikacji na stronie szkoły i w serwisach opinii."),
                "opinion": extra.get("opinion", "Sprawdzić opinie rodziców i uczniów dla konkretnych rozszerzeń."),
                "why": make_why(record, top),
                "watch": make_watch(record, top),
                "gallery": [gallery_url(name) for name in gallery_files],
                "gallerySources": [extra.get("commons", "") or image_search_url(school), image_search_url(school)],
                "opinionSources": [opinion_search_url(school)],
            }
        )
        records.append(record)
    return sorted(records, key=lambda item: item["bestScore"], reverse=True)


def make_why(record: dict, top: dict) -> str:
    p = top["P - politechnika 0-100"]
    m = top["M - medycyna 0-100"]
    profile = top["Klasa/kierunek"]
    if m >= 80 and p >= 80:
        return f"Najmocniejsza klasa ({profile}) daje jednocześnie medyczny i techniczny zapas: biologia/chemia lub chemia/fizyka z matematyką."
    if p >= 90 and m >= 40:
        return f"Najmocniejsza klasa ({profile}) jest bardzo dobra pod politechnikę i zostawia ścieżkę okołomedyczną, np. inżynierię biomedyczną."
    if p >= 90:
        return f"Najmocniejsza klasa ({profile}) jest przede wszystkim politechniczna: matematyka, fizyka, informatyka lub technika."
    if m >= 90:
        return f"Najmocniejsza klasa ({profile}) jest przede wszystkim medyczna: biologia i chemia jako główne rozszerzenia."
    return f"Najmocniejsza klasa ({profile}) jest sensowna jako uzupełnienie listy preferencji."


def make_watch(record: dict, top: dict) -> str:
    risks = []
    if top["Szansa przy 168 pkt"] in ("ambitna", "bardzo ambitna"):
        risks.append("wysokie ryzyko punktowe")
    if top["Dojazd z Kobyłki"] == "słaby":
        risks.append("codzienny dojazd")
    if top["Egzamin dodatkowy z języka angielskiego"].startswith("TAK"):
        risks.append("dodatkowy sprawdzian językowy")
    if not risks:
        return "Główne ryzyko jest umiarkowane; mimo to warto sprawdzić atmosferę w konkretnej klasie i nauczycieli rozszerzeń."
    return "Najważniejsze do sprawdzenia: " + ", ".join(risks) + "."


def write_site(records: list[dict]) -> None:
    SITE_DIR.mkdir(exist_ok=True)
    (SITE_DIR / "data.js").write_text(
        "window.SCHOOLS = " + json.dumps(records, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    (SITE_DIR / "index.html").write_text(INDEX_HTML, encoding="utf-8")
    (SITE_DIR / "styles.css").write_text(STYLES_CSS, encoding="utf-8")
    (SITE_DIR / "app.js").write_text(APP_JS, encoding="utf-8")


INDEX_HTML = """<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Szkoły dla syna | Warszawa</title>
  <link rel="stylesheet" href="styles.css" />
</head>
<body>
  <header class="topbar">
    <div>
      <p class="eyebrow">Warszawa · licea i technika · preferencje rekrutacyjne</p>
      <h1>Szkoły dla syna</h1>
      <p class="lead">Jednolita prezentacja szkół z rankingiem atrakcyjności, ryzykiem punktowym, galeriami i argumentem „dlaczego ta szkoła”.</p>
    </div>
    <div class="summary">
      <strong>Cel</strong>
      <span>politechnika albo medycyna</span>
      <strong>Założenie</strong>
      <span>ok. 168 pkt</span>
    </div>
  </header>

  <main>
    <section class="controls">
      <input id="search" type="search" placeholder="Szukaj szkoły, dzielnicy, profilu..." />
      <select id="district"></select>
      <select id="risk">
        <option value="">Wszystkie ryzyka</option>
        <option>bardzo ambitna</option>
        <option>ambitna</option>
        <option>na granicy</option>
        <option>realna</option>
        <option>bezpieczna</option>
      </select>
    </section>
    <section id="schoolGrid" class="grid"></section>
  </main>

  <footer>
    <p>Źródła: lokalny arkusz progów 2023-2025, oficjalne strony szkół, Wikimedia Commons, wyszukiwarki opinii/forum. Opinie internetowe są wskazówką do dalszej weryfikacji, nie faktem rankingowym.</p>
  </footer>
  <script src="data.js"></script>
  <script src="app.js"></script>
</body>
</html>
"""


STYLES_CSS = """
:root{--bg:#f7f8f5;--ink:#17201b;--muted:#66736b;--line:#dfe5de;--card:#fff;--accent:#256b54;--accent2:#b45309;--blue:#244f7a}
*{box-sizing:border-box}body{margin:0;font-family:Inter,ui-sans-serif,system-ui,-apple-system,Segoe UI,sans-serif;background:var(--bg);color:var(--ink)}
.topbar{display:grid;grid-template-columns:minmax(0,1fr) 260px;gap:32px;padding:48px 56px 28px;border-bottom:1px solid var(--line);background:#fff}
.eyebrow{margin:0 0 10px;text-transform:uppercase;font-size:12px;letter-spacing:.08em;color:var(--accent);font-weight:700}h1{margin:0;font-size:48px;line-height:1}h2{font-size:22px;margin:0}.lead{max-width:820px;color:var(--muted);font-size:18px;line-height:1.5}.summary{border:1px solid var(--line);border-radius:8px;padding:18px;background:#fbfcfa;display:grid;gap:6px;align-content:start}.summary strong{font-size:12px;text-transform:uppercase;color:var(--muted)}.summary span{font-size:16px}
main{padding:26px 56px 56px}.controls{display:grid;grid-template-columns:minmax(260px,1fr) 220px 200px;gap:12px;margin-bottom:22px}input,select{height:44px;border:1px solid var(--line);border-radius:8px;background:#fff;padding:0 14px;font:inherit;color:var(--ink)}
.grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px}.card{background:var(--card);border:1px solid var(--line);border-radius:8px;overflow:hidden;display:grid;grid-template-columns:220px minmax(0,1fr);min-height:360px}.media{background:#e8ece8;display:grid;grid-template-rows:1fr 1fr 1fr;gap:2px}.media img{width:100%;height:100%;object-fit:cover;display:block}.fallback{display:grid;place-items:center;text-align:center;padding:18px;color:#54615a;background:linear-gradient(135deg,#edf2ee,#dfe8e2);font-weight:700}
.content{padding:18px 18px 16px}.meta{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0 14px}.pill{border:1px solid var(--line);border-radius:999px;padding:5px 9px;font-size:12px;color:#39443d;background:#f7faf7}.score{color:#fff;background:var(--accent);border-color:var(--accent)}.risk{background:#fff7ed;border-color:#fed7aa;color:#9a3412}.matura{display:grid;grid-template-columns:auto 1fr;gap:2px 10px;align-items:baseline;border:1px solid var(--line);border-radius:8px;padding:10px 12px;background:#fbfcfa;margin-bottom:12px}.matura span{font-size:12px;text-transform:uppercase;color:var(--muted);font-weight:800}.matura strong{font-size:22px;color:var(--accent)}.matura small{grid-column:1/-1;color:var(--muted)}.section{border-top:1px solid var(--line);padding-top:12px;margin-top:12px}.section h3{font-size:12px;text-transform:uppercase;letter-spacing:.07em;color:var(--muted);margin:0 0 6px}.section p{margin:0;color:#344039;line-height:1.45}.classes{display:grid;gap:8px;margin-top:8px}.classrow{display:grid;gap:5px;font-size:13px;padding:9px 10px;background:#f7f8f5;border-radius:6px}.classhead{display:grid;grid-template-columns:1fr auto;gap:8px}.classmeta{color:var(--muted);line-height:1.35}.links{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}.links a{color:var(--blue);text-decoration:none;font-weight:650;font-size:13px}.links a:hover{text-decoration:underline}footer{padding:22px 56px 40px;color:var(--muted);border-top:1px solid var(--line)}
@media(max-width:980px){.topbar{grid-template-columns:1fr;padding:32px 20px}.controls{grid-template-columns:1fr}.grid{grid-template-columns:1fr}main{padding:20px}.card{grid-template-columns:1fr}.media{height:260px;grid-template-columns:1fr 1fr 1fr;grid-template-rows:1fr}h1{font-size:36px}}
"""


APP_JS = """
const grid = document.getElementById('schoolGrid');
const search = document.getElementById('search');
const district = document.getElementById('district');
const risk = document.getElementById('risk');

const districts = [...new Set(window.SCHOOLS.map(s => s.district))].sort();
district.innerHTML = '<option value="">Wszystkie dzielnice</option>' + districts.map(d => `<option>${d}</option>`).join('');

function render(){
  const q = search.value.toLowerCase();
  const d = district.value;
  const r = risk.value;
  const data = window.SCHOOLS.filter(s => {
    const text = [s.school, s.district, s.summary, ...s.classes.map(c => c['Klasa/kierunek'])].join(' ').toLowerCase();
    return (!q || text.includes(q)) && (!d || s.district === d) && (!r || s.bestRisk === r || s.classes.some(c => c['Szansa przy 168 pkt'] === r));
  });
  grid.innerHTML = data.map(card).join('');
}

function card(s){
  const top = s.classes[0];
  const gallery = s.gallery.length ? s.gallery.slice(0,3).map(src => `<img src="${src}" loading="lazy" alt="${s.school}">`).join('') : `<div class="fallback">${s.school}<br><small>linki do zdjęć w źródłach</small></div><div class="fallback">galeria</div><div class="fallback">zdjęcia</div>`;
  const matura = `<div class="matura">
    <span>${top['Matura międzynarodowa IB'] === 'TAK' ? 'IB DP' : 'matura polska'}</span>
    <strong>${top['Wskaźnik maturalny 2026'] || 'brak danych'}</strong>
    <small>ranking 2026: ${top['Ranking maturalny 2026'] || 'brak danych'} · E ${top['E - matura szkoły 0-100']}</small>
  </div>`;
  const classes = s.classes.slice(0,3).map(c => `<div class="classrow">
    <div class="classhead"><span>${c['Klasa/kierunek']}</span><strong>${c['Atrakcyjność 0-100']}</strong></div>
    <div class="classmeta">Punktowane: ${c['Przedmioty oceniane/punktowane'] || 'sprawdzić w rekrutacji'}</div>
    <div class="classmeta">Progi: 2024 ${c['Próg 2024'] || '-'}, 2025 ${c['Próg 2025'] || '-'}, średnia ${c['Średnia progów'] || '-'}</div>
    <div class="classmeta">Egzamin z angielskiego: ${c['Egzamin dodatkowy z języka angielskiego'] || 'nie'}</div>
  </div>`).join('');
  const links = [
    s.official && `<a href="${s.official}" target="_blank">strona szkoły</a>`,
    s.commons && `<a href="${s.commons}" target="_blank">Wikimedia</a>`,
    `<a href="${s.gallerySources[1]}" target="_blank">więcej zdjęć</a>`,
    `<a href="${s.opinionSources[0]}" target="_blank">opinie i fora</a>`
  ].filter(Boolean).join('');
  return `<article class="card">
    <div class="media">${gallery}</div>
    <div class="content">
      <h2>${s.school}</h2>
      <div class="meta"><span class="pill">${s.district}</span><span class="pill score">atrakcyjność ${s.bestScore}</span><span class="pill risk">${s.bestRisk}</span></div>
      ${matura}
      <div class="section"><h3>Opis szkoły</h3><p>${s.summary}</p></div>
      <div class="section"><h3>Dlaczego dla syna</h3><p>${s.why}</p></div>
      <div class="section"><h3>Na co uważać</h3><p>${s.watch}</p></div>
      <div class="section"><h3>Najlepsze klasy z listy</h3><div class="classes">${classes}</div></div>
      <div class="section"><h3>Opinie z internetu</h3><p>${s.opinion}</p><div class="links">${links}</div></div>
    </div>
  </article>`;
}

[search,district,risk].forEach(el => el.addEventListener('input', render));
render();
"""


def main() -> None:
    records = build_school_records(rows_from_workbook())
    write_site(records)
    print(SITE_DIR / "index.html")


if __name__ == "__main__":
    main()
