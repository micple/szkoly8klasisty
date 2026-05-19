"""
Liczy czasy dojazdu rano (poniedziałek 07:30, Warszawa) z Kobyłki do każdej szkoły
przez Google Routes API i dopisuje arkusz 'Czasy dojazdu rano' do
outputs/wybor_szkol_warszawa.xlsx.

Klucz API jest referer-restricted (localhost) — wysyłam ręcznie Referer header,
co Routes API akceptuje (Geocoding REST już nie — ale tego nie potrzebujemy).

Wymaga: requests, openpyxl. Python 3.9+.

Uruchomienie:  python scripts/add-commute-to-xlsx.py
"""

from __future__ import annotations
import io
import json
import re
import sys
import time

# Wymuś UTF-8 na stdout (Windows ma cp1250 domyślnie — psuje Polish chars).
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


ROOT = Path(__file__).resolve().parent.parent
DATA_JS = ROOT / "site" / "data.js"
CONFIG_LOCAL_JS = ROOT / "site-v2" / "config.local.js"
XLSX_PATH = ROOT / "outputs" / "wybor_szkol_warszawa.xlsx"

KOBYLKA = "Kobyłka, Polska"
REFERER = "http://localhost:8000/site-v2/index.html"
SHEET_NAME = "Czasy dojazdu rano"
DEPARTURE_HOUR = 7
DEPARTURE_MIN = 30
THROTTLE_MS = 200


def load_api_key() -> str:
    txt = CONFIG_LOCAL_JS.read_text(encoding="utf-8")
    m = re.search(r"window\.GMAPS_API_KEY\s*=\s*['\"]([^'\"]+)['\"]", txt)
    if not m:
        sys.exit("Nie znaleziono klucza w config.local.js")
    return m.group(1)


def load_schools() -> list[dict[str, Any]]:
    txt = DATA_JS.read_text(encoding="utf-8")
    start = txt.find("[")
    end = txt.rfind("]")
    if start < 0 or end < 0:
        sys.exit("Nieprawidłowy format data.js")
    return json.loads(txt[start : end + 1])


def next_monday_iso(hour: int, minute: int) -> str:
    """Returns next Monday at HH:MM in Warsaw, formatted as RFC3339 UTC."""
    tz = ZoneInfo("Europe/Warsaw")
    today = datetime.now(tz).date()
    days_ahead = (7 - today.weekday()) % 7  # Mon=0
    if days_ahead == 0:
        days_ahead = 7
    target_date = today + timedelta(days=days_ahead)
    dt = datetime(target_date.year, target_date.month, target_date.day, hour, minute, tzinfo=tz)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def fetch_commute(api_key: str, lat: float, lng: float, departure_iso: str) -> dict[str, Any]:
    payload = {
        "origin": {"address": KOBYLKA},
        "destination": {"location": {"latLng": {"latitude": lat, "longitude": lng}}},
        "travelMode": "TRANSIT",
        "departureTime": departure_iso,
        "languageCode": "pl",
        "regionCode": "PL",
    }
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": "routes.duration,routes.distanceMeters,routes.legs.steps.travelMode,routes.legs.steps.transitDetails.transitLine.nameShort,routes.legs.steps.transitDetails.stopDetails.arrivalTime,routes.legs.steps.transitDetails.stopDetails.departureTime",
        "Referer": REFERER,
    }
    r = requests.post(
        "https://routes.googleapis.com/directions/v2:computeRoutes",
        headers=headers,
        json=payload,
        timeout=30,
    )
    if r.status_code != 200:
        return {"error": f"HTTP {r.status_code}: {r.text[:200]}"}
    data = r.json()
    routes = data.get("routes") or []
    if not routes:
        return {"error": "brak trasy w odpowiedzi"}
    route = routes[0]
    duration_str = route.get("duration") or "0s"
    duration_sec = int(duration_str.rstrip("s"))
    distance_m = route.get("distanceMeters") or 0
    transit_steps = []
    departure_time = None
    arrival_time = None
    for leg in route.get("legs", []):
        for step in leg.get("steps", []):
            if step.get("travelMode") == "TRANSIT":
                td = step.get("transitDetails", {}) or {}
                line = td.get("transitLine", {}).get("nameShort") or ""
                transit_steps.append(line)
                stop = td.get("stopDetails", {}) or {}
                if departure_time is None:
                    departure_time = stop.get("departureTime")
                if stop.get("arrivalTime"):
                    arrival_time = stop.get("arrivalTime")
    transfers = max(0, len(transit_steps) - 1)
    return {
        "duration_sec": duration_sec,
        "duration_min": round(duration_sec / 60),
        "duration_text": format_duration(duration_sec),
        "distance_km": round(distance_m / 1000, 1),
        "transfers": transfers,
        "lines": ", ".join(transit_steps),
        "departure_iso": departure_time,
        "arrival_iso": arrival_time,
    }


def format_duration(sec: int) -> str:
    h, rem = divmod(sec, 3600)
    m = round(rem / 60)
    if h:
        return f"{h}h {m}min"
    return f"{m}min"


def format_warsaw_time(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.astimezone(ZoneInfo("Europe/Warsaw")).strftime("%H:%M")
    except Exception:
        return ""


def main() -> int:
    api_key = load_api_key()
    schools = load_schools()
    departure_iso = next_monday_iso(DEPARTURE_HOUR, DEPARTURE_MIN)
    print(f"Departure: {departure_iso} (pon. {DEPARTURE_HOUR:02d}:{DEPARTURE_MIN:02d} Warszawa)")
    print(f"Szkół do policzenia: {len(schools)}\n")

    rows: list[dict[str, Any]] = []
    for i, s in enumerate(schools, start=1):
        name = s.get("school", "")
        district = s.get("district", "")
        lat, lng = s.get("lat"), s.get("lng")
        if not isinstance(lat, (int, float)) or not isinstance(lng, (int, float)):
            print(f"[{i}/{len(schools)}] ⊘ {name} — brak lat/lng")
            rows.append({"name": name, "district": district, "error": "brak współrzędnych"})
            continue
        try:
            res = fetch_commute(api_key, float(lat), float(lng), departure_iso)
        except Exception as e:
            res = {"error": str(e)}
        if "error" in res:
            print(f"[{i}/{len(schools)}] ✗ {name} — {res['error']}")
            rows.append({"name": name, "district": district, "error": res["error"]})
        else:
            print(
                f"[{i}/{len(schools)}] ✓ {name}: {res['duration_text']} · "
                f"{res['transfers']} przesiad. · {res['distance_km']} km"
            )
            rows.append({"name": name, "district": district, **res})
        time.sleep(THROTTLE_MS / 1000)

    # Statystyki
    ok = [r for r in rows if "duration_sec" in r]
    avg_min = sum(r["duration_min"] for r in ok) / len(ok) if ok else 0
    print(
        f"\nKoniec: {len(ok)} policzonych, {len(rows) - len(ok)} błędów. "
        f"Średnia: {avg_min:.0f} min."
    )

    write_xlsx(rows, departure_iso)
    return 0


# Zamieniamy istniejącą kolumnę "Dojazd z Kobyłki" (jakościowe "dobry/średni/słaby")
# na realny czas dojazdu rano w minutach. Header pozostaje "Dojazd z Kobyłki (rano)".
TARGET_COL_HEADER = "Dojazd z Kobyłki"
NEW_HEADER = "Dojazd z Kobyłki (min, rano)"


def write_xlsx(rows: list[dict[str, Any]], departure_iso: str) -> None:
    by_school = {r["name"]: r for r in rows if "duration_sec" in r}
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Usuń starą osobną zakładkę jeśli jest (user chce wszystko inline)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"\nUsunięto starą zakładkę '{SHEET_NAME}' — dane będą inline w głównych arkuszach")

    # Border + fill helpers
    thin_border = Border(
        left=Side(style="thin", color="DCD3BD"),
        right=Side(style="thin", color="DCD3BD"),
        top=Side(style="thin", color="DCD3BD"),
        bottom=Side(style="thin", color="DCD3BD"),
    )

    def time_fill(minutes: Any) -> PatternFill | None:
        if not isinstance(minutes, int):
            return None
        if minutes < 30:   return PatternFill("solid", fgColor="DDEFDF")
        if minutes < 50:   return PatternFill("solid", fgColor="E7F0D2")
        if minutes < 70:   return PatternFill("solid", fgColor="F8ECC9")
        return PatternFill("solid", fgColor="FBE6D0")

    # Dla każdego głównego arkusza: ZAMIEŃ kolumnę "Dojazd z Kobyłki" na realne czasy.
    for sheet_name in ("Od pierwszego wyboru", "Rekomendacja"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [(c.value or "").strip() if c.value else "" for c in ws[1]]

        # Znajdź kolumnę "Szkoła" + "Dojazd z Kobyłki"
        try:
            school_col_idx = headers.index("Szkoła") + 1
        except ValueError:
            print(f"⚠ {sheet_name}: brak kolumny 'Szkoła', pomijam")
            continue

        target_col_idx = None
        for i, h in enumerate(headers):
            if h.startswith(TARGET_COL_HEADER):
                target_col_idx = i + 1
                break
        if target_col_idx is None:
            print(f"⚠ {sheet_name}: brak kolumny '{TARGET_COL_HEADER}', pomijam")
            continue

        # Usuń ewentualną poprzednio doklejoną kolumnę "Czas dojazdu rano" na końcu
        for i in range(len(headers) - 1, -1, -1):
            if headers[i] == "Czas dojazdu rano":
                ws.delete_cols(i + 1, 1)
                print(f"  ⊘ {sheet_name}: usuwam duplikat 'Czas dojazdu rano' z końca (kol {i+1})")

        # Zmień nagłówek + styl
        hcell = ws.cell(row=1, column=target_col_idx, value=NEW_HEADER)
        hcell.font = Font(bold=True, color="FFFFFF")
        hcell.fill = PatternFill("solid", fgColor="1F1B16")
        hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hcell.border = thin_border

        # Wypełnij wartości dla każdego wiersza
        filled = 0
        for row_idx in range(2, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=school_col_idx).value
            if not name_cell:
                continue
            r = by_school.get(str(name_cell).strip())
            cell = ws.cell(row=row_idx, column=target_col_idx)
            if not r:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
            else:
                m = r.get("duration_min")
                cell.value = int(m) if m is not None else None
                cell.number_format = "0"
                fill = time_fill(m)
                if fill:
                    cell.fill = fill
                filled += 1
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center")

        ws.column_dimensions[openpyxl.utils.get_column_letter(target_col_idx)].width = 18
        print(f"✓ {sheet_name}: kolumna {target_col_idx} ({NEW_HEADER}) — {filled} wartości")

    wb.save(XLSX_PATH)
    print(f"\nZapisano: {XLSX_PATH}")


if __name__ == "__main__":
    sys.exit(main())
