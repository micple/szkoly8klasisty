from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "wyniki lata przeszle"
OUTPUT_DIR = ROOT / "outputs"
OUTPUT_FILE = OUTPUT_DIR / "wybor_szkol_warszawa.xlsx"
ASSUMED_POINTS = 168.0


def repair_text(text: str) -> str:
    try:
        return text.encode("cp1250").decode("utf-8")
    except UnicodeError:
        return text


def clean_cell(value: str) -> str:
    value = repair_text(value)
    value = re.sub(r"<br\s*/?>", "<br>", value)
    value = value.replace("&nbsp;", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def parse_float(value: str) -> float | None:
    value = clean_cell(value).replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", value)
    if not match:
        return None
    return float(match.group(0))


def parse_int(value: str) -> int | None:
    value = clean_cell(value).replace("=", "").strip()
    if value in {"I", "ı", "l", "П", "Đź"}:
        return 1
    match = re.search(r"\d+", value)
    if not match:
        return None
    return int(match.group(0))


def split_multi(value: str) -> list[str]:
    parts = [clean_cell(part) for part in value.split("<br>")]
    return [part for part in parts if part]


def unsplit_cell(value: str) -> str:
    return clean_cell(value).replace("<br>", " ")


def canonical_class(value: str) -> str:
    value = unsplit_cell(value)
    markers = list(re.finditer(r"\[[A-ZŁŚIWMSo0-9-]+\]", value))
    if markers:
        value = value[markers[-1].start() :]
    return re.sub(r"\s+", " ", value).strip()


def year_from_name(path: Path) -> int:
    match = re.search(r"20\d{2}", path.name)
    if not match:
        raise ValueError(f"Cannot infer year from {path.name}")
    return int(match.group(0))


@dataclass
class ThresholdRow:
    year: int
    district: str
    school: str
    symbol: str
    class_name: str
    threshold: float | None
    source_file: str


@dataclass
class MaturaRow:
    school: str
    school_type: str
    city: str
    rank_2026: int | None
    rank_2025: int | None
    rank_2024: int | None
    overall: float | None
    required: float | None
    additional: float | None
    source_file: str
    source_line: int


def parse_markdown_tables() -> list[ThresholdRow]:
    rows: list[ThresholdRow] = []
    for path in sorted(SOURCE_DIR.glob("*.md")):
        year = year_from_name(path)
        text = path.read_text(encoding="utf-8", errors="replace")
        for line in text.splitlines():
            line = line.strip()
            if not line.startswith("|") or "---" in line:
                continue
            cells = [clean_cell(cell) for cell in line.strip("|").split("|")]
            if len(cells) < 5 or "Dzielnica" in cells[0]:
                continue
            if cells[0].lower() == "dzielnica":
                continue
            if year == 2024 and len(cells) >= 6:
                district_parts = [unsplit_cell(cells[0])]
                school_parts = [unsplit_cell(cells[2])]
                symbol_parts = [unsplit_cell(cells[4]).split(" ", 1)[0]]
                class_parts = [canonical_class(cells[4])]
                threshold_parts = [unsplit_cell(cells[5])]
            elif year == 2023 and len(cells) >= 6:
                district_parts = [unsplit_cell(cells[0])]
                school_parts = [unsplit_cell(cells[1])]
                symbol_parts = [unsplit_cell(cells[2])]
                class_parts = [canonical_class(cells[3])]
                threshold_parts = [unsplit_cell(cells[4])]
            else:
                district_parts = split_multi(cells[0])
                school_parts = split_multi(cells[1])
                symbol_parts = split_multi(cells[2])
                class_parts = [canonical_class(part) for part in split_multi(cells[3])]
                threshold_parts = split_multi(cells[4])

            count = max(
                len(district_parts),
                len(school_parts),
                len(symbol_parts),
                len(class_parts),
                len(threshold_parts),
            )
            for index in range(count):
                district = district_parts[index] if index < len(district_parts) else (district_parts[-1] if district_parts else "")
                school = school_parts[index] if index < len(school_parts) else (school_parts[-1] if school_parts else "")
                symbol = symbol_parts[index] if index < len(symbol_parts) else ""
                class_name = class_parts[index] if index < len(class_parts) else ""
                threshold = parse_float(threshold_parts[index]) if index < len(threshold_parts) else None
                if not school or not class_name:
                    continue
                rows.append(
                    ThresholdRow(
                        year=year,
                        district=district,
                        school=school,
                        symbol=symbol,
                        class_name=class_name,
                        threshold=threshold,
                        source_file=path.name,
                    )
                )
    return rows


def parse_matura_rankings() -> list[MaturaRow]:
    rows: list[MaturaRow] = []
    files = [
        ("LO", SOURCE_DIR / "ranking-licea-matura-2026.md"),
        ("Technikum", SOURCE_DIR / "ranking-technika-matura-2026.md"),
    ]
    for school_type, path in files:
        if not path.exists():
            continue
        for line_no, line in enumerate(path.read_text(encoding="utf-8", errors="replace").splitlines(), start=1):
            stripped = line.strip()
            if stripped.startswith("|") and "---" not in stripped:
                cells = [clean_cell(cell) for cell in stripped.strip("|").split("|")]
                if "Warszawa" not in cells:
                    continue
                city_index = cells.index("Warszawa")
                if city_index < 2:
                    continue
                school = cells[city_index - 1]
                rank_2026 = parse_int(cells[0])
                rank_2025 = parse_int(cells[4]) if len(cells) >= 16 else (parse_int(cells[3]) if len(cells) >= 8 else None)
                rank_2024 = parse_int(cells[5]) if len(cells) >= 16 else (parse_int(cells[4]) if len(cells) >= 8 else None)
                if len(cells) >= 16:
                    overall = parse_float(cells[7])
                    required = parse_float(cells[8])
                    additional = parse_float(cells[9])
                elif len(cells) >= 8:
                    overall = parse_float(cells[5])
                    required = parse_float(cells[6])
                    additional = parse_float(cells[7])
                else:
                    continue
                if not school or overall is None:
                    continue
                rows.append(
                    MaturaRow(
                        school=school,
                        school_type=school_type,
                        city="Warszawa",
                        rank_2026=rank_2026,
                        rank_2025=rank_2025,
                        rank_2024=rank_2024,
                        overall=overall,
                        required=required,
                        additional=additional,
                        source_file=path.name,
                        source_line=line_no,
                    )
                )

            if school_type == "LO" and "XII LO im. Henryka Sienkiewicza Warszawa" in clean_cell(line):
                text = clean_cell(line)
                match = re.search(
                    r"XII LO im\. Henryka Sienkiewicza Warszawa\s+(\d+)\s+(\d+)\s+(\d+[,.]\d+)\s+(\d+[,.]\d+)\s+(\d+[,.]\d+)",
                    text,
                )
                if match:
                    rows.append(
                        MaturaRow(
                            school="XII LO im. Henryka Sienkiewicza",
                            school_type="LO",
                            city="Warszawa",
                            rank_2026=None,
                            rank_2025=parse_int(match.group(1)),
                            rank_2024=parse_int(match.group(2)),
                            overall=parse_float(match.group(3)),
                            required=parse_float(match.group(4)),
                            additional=parse_float(match.group(5)),
                            source_file=path.name,
                            source_line=line_no,
                        )
                    )
    return rows


def norm(value: str) -> str:
    value = repair_text(value).lower()
    replacements = {
        "ą": "a",
        "ć": "c",
        "ę": "e",
        "ł": "l",
        "ń": "n",
        "ó": "o",
        "ś": "s",
        "ź": "z",
        "ż": "z",
        "„": "",
        "”": "",
        '"': "",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    value = value.replace("nr i", "nr 1")
    return re.sub(r"\s+", " ", value)


def match_candidate(rows: list[ThresholdRow], school_tokens: list[str], class_tokens: list[str]) -> list[ThresholdRow]:
    school_tokens_norm = [[norm(option) for option in token.split("|")] for token in school_tokens]
    class_tokens_norm = [norm(token) for token in class_tokens]
    matches = []
    for row in rows:
        school = norm(row.school)
        class_name = norm(row.class_name)
        if all(any(option in school for option in token_group) for token_group in school_tokens_norm) and all(token in class_name for token in class_tokens_norm):
            matches.append(row)
    return matches


def match_matura(rows: list[MaturaRow], school_tokens: list[str]) -> MaturaRow | None:
    school_tokens_norm = [[norm(option) for option in token.split("|")] for token in school_tokens]
    matches = []
    for row in rows:
        if norm(row.city) != "warszawa":
            continue
        school = norm(row.school)
        if all(any(option in school for option in token_group) for token_group in school_tokens_norm):
            matches.append(row)
    if not matches:
        return None
    return max(matches, key=lambda item: item.overall or 0)


def chance_label(delta_2025: float | None) -> str:
    if delta_2025 is None:
        return "brak danych"
    if delta_2025 >= 10:
        return "bezpieczna"
    if delta_2025 >= 3:
        return "realna"
    if delta_2025 >= -3:
        return "na granicy"
    if delta_2025 >= -8:
        return "ambitna"
    return "bardzo ambitna"


def fit_score(profile: str, school_type: str) -> tuple[int, int]:
    p = norm(profile)
    if "ib" in p or "international baccalaureate" in p or "miedzynarod" in p:
        return 4, 4
    if "mechatronik" in p or "automatyk" in p:
        return 5, 2
    if "elektronik" in p or "teleinformatyk" in p:
        return 5, 1
    if "programista" in p or "informatyk" in p:
        return 5, 1
    if "technik" in p:
        return 3, 0

    has_mat = "mat" in p
    has_fiz = "fiz" in p
    has_inf = "inf" in p
    has_chem = "chem" in p
    has_biol = "biol" in p
    has_geogr = "geogr" in p

    if has_fiz and has_inf and has_mat:
        poly = 5
    elif has_fiz and has_mat:
        poly = 5
    elif has_inf and has_mat:
        poly = 5
    elif has_chem and has_fiz and has_mat:
        poly = 5
    elif has_biol and has_chem and has_mat:
        poly = 4
    elif has_geogr and has_mat:
        poly = 3
    elif has_chem and has_mat:
        poly = 3
    elif has_mat:
        poly = 3
    elif has_chem:
        poly = 2
    else:
        poly = 1

    if has_biol and has_chem and has_mat:
        med = 5
    elif has_biol and has_chem:
        med = 5
    elif has_chem and has_fiz and has_mat:
        med = 4
    elif has_fiz and has_inf and has_mat:
        med = 2
    elif has_fiz and has_mat:
        med = 1
    elif has_inf and has_mat:
        med = 1
    elif has_biol and has_mat:
        med = 3
    elif has_chem and has_mat:
        med = 3
    elif has_biol or has_chem:
        med = 2
    else:
        med = 0

    return poly, med


def admission_score(threshold_2025: float | None) -> int:
    if threshold_2025 is None:
        return 40
    if threshold_2025 <= 158:
        return 100
    if threshold_2025 <= 165:
        return 85
    if threshold_2025 <= 168:
        return 70
    if threshold_2025 <= 171:
        return 55
    if threshold_2025 <= 176:
        return 35
    return 15


def class_strength_score(threshold_2025: float | None) -> int:
    if threshold_2025 is None:
        return 40
    if threshold_2025 >= 178:
        return 100
    if threshold_2025 >= 172:
        return 90
    if threshold_2025 >= 168:
        return 80
    if threshold_2025 >= 162:
        return 65
    if threshold_2025 >= 155:
        return 50
    return 35


def commute_score(commute: str) -> int:
    return {"dobry": 100, "średni": 65, "słaby": 30}.get(commute, 50)


def stability_score(values: list[float | None]) -> int:
    nums = [value for value in values if value is not None]
    if len(nums) <= 1:
        return 40
    spread = max(nums) - min(nums)
    if spread <= 3:
        return 100
    if spread <= 6:
        return 75
    if spread <= 10:
        return 50
    return 25


def matura_quality_score(row: MaturaRow | None, is_ib: bool = False) -> float:
    if row is None or row.overall is None:
        if is_ib:
            return 70.0
        return 45.0
    required = row.required if row.required is not None else row.overall
    additional = row.additional if row.additional is not None else row.overall
    return round(0.60 * row.overall + 0.20 * required + 0.20 * additional, 2)


def final_score(
    poly: int,
    med: int,
    threshold_2025: float | None,
    commute: str,
    values: list[float | None],
    matura: MaturaRow | None,
    is_ib: bool = False,
) -> tuple[int, int, int, int, int, int, float, float]:
    p = poly * 20
    m = med * 20
    q = class_strength_score(threshold_2025)
    d = commute_score(commute)
    s = stability_score(values)
    e = matura_quality_score(matura, is_ib)
    fit = round(0.75 * max(p, m) + 0.25 * min(p, m), 2)
    total = round(0.30 * fit + 0.25 * e + 0.20 * q + 0.15 * d + 0.10 * s, 2)
    return p, m, fit, q, d, s, e, total


SUBJECT_MAP = {
    "ang": "język angielski",
    "biz": "biznes i zarządzanie",
    "biol": "biologia",
    "chem": "chemia",
    "fiz": "fizyka",
    "fra": "język francuski",
    "geogr": "geografia",
    "hist": "historia",
    "hisz": "język hiszpański",
    "hiszp": "język hiszpański",
    "inf": "informatyka",
    "mat": "matematyka",
    "niem": "język niemiecki",
    "pol": "język polski",
    "ros": "język rosyjski",
    "wlo": "język włoski",
    "wos": "wiedza o społeczeństwie",
}


def class_marker(profile: str) -> str:
    match = re.search(r"\[([^\]]+)\]", profile)
    return match.group(1) if match else ""


def profile_codes(profile: str) -> list[str]:
    text = norm(profile)
    match = re.search(r"\]\s*([a-z.-]+(?:-[a-z.]+){1,3})", text)
    if not match:
        match = re.search(r"\b([a-z]+(?:-[a-z]+){1,3})\b", text)
    if not match:
        return []
    return [part.strip(".") for part in match.group(1).split("-") if part.strip(".")]


def profile_subjects(profile: str) -> str:
    if "ib" in norm(profile) or "miedzynarod" in norm(profile):
        return "IB DP: wybór 6 przedmiotów, zwykle 3 na Higher Level; dla celów syna kluczowe HL Math, Physics/Chemistry/Biology"
    if "technik" in norm(profile):
        match = re.search(r"Technik[^()]+", profile, flags=re.IGNORECASE)
        zawod = re.sub(r"\s+", " ", match.group(0)).strip() if match else "kierunek technikum"
        langs_match = re.search(r"\(([^)]+)\)", profile)
        langs = ""
        if langs_match:
            lang_codes = [part.strip("* ") for part in re.split(r"[-,]", langs_match.group(1)) if part.strip("* ")]
            lang_names = [SUBJECT_MAP.get(norm(code), code) for code in lang_codes]
            langs = f"; języki: {', '.join(lang_names)}"
        return f"{zawod}{langs}"
    codes = profile_codes(profile)
    subjects = [SUBJECT_MAP.get(code, code) for code in codes]
    if subjects:
        return ", ".join(subjects)
    return "brak jednoznacznego profilu w nazwie oddziału"


def scored_subjects(profile: str) -> str:
    subjects = profile_subjects(profile)
    if subjects.startswith("IB DP"):
        return "rekrutacja własna/oddział międzynarodowy; kluczowy angielski i późniejszy wybór przedmiotów HL"
    if subjects.lower().startswith("technik"):
        return "język polski, matematyka oraz przedmioty wskazane w ofercie oddziału technikum"
    if subjects.startswith("brak"):
        return "język polski, matematyka oraz przedmioty wskazane w ofercie oddziału"
    return f"język polski, matematyka oraz profil: {subjects}"


def language_exam(profile: str) -> str:
    if "ib" in norm(profile) or "miedzynarod" in norm(profile):
        return "TAK - ścieżka IB/pre-IB; szkoła zwykle wymaga bardzo dobrego angielskiego i/lub egzaminu/rozmowy"
    marker = class_marker(profile)
    if marker == "D":
        return "TAK - oddział dwujęzyczny; zwykle wymagany sprawdzian kompetencji językowych"
    if marker == "M":
        return "TAK - oddział międzynarodowy; zwykle wymagany sprawdzian/predyspozycje językowe"
    return "nie wynika z oznaczenia oddziału"


def priority_score(row: dict) -> float:
    base = row["Dopasowanie politechnika"] * 1.4 + row["Dopasowanie medycyna"] * 0.9
    if row["Typ"] == "Technikum":
        base += 0.5
    if row["Dojazd z Kobyłki"] == "dobry":
        base += 1.0
    elif row["Dojazd z Kobyłki"] == "średni":
        base += 0.3
    if row["Szansa przy 168 pkt"] == "bezpieczna":
        base += 1.2
    elif row["Szansa przy 168 pkt"] == "realna":
        base += 1.0
    elif row["Szansa przy 168 pkt"] == "na granicy":
        base += 0.4
    elif row["Szansa przy 168 pkt"] == "ambitna":
        base -= 0.4
    else:
        base -= 1.2
    return round(base, 2)


CANDIDATES = [
    {
        "source_order": 1,
        "school": "Technikum Mechatroniczne nr 1",
        "type": "Technikum",
        "class": "Technik mechatronik (ang-hisz)",
        "school_tokens": ["Technikum Mechatroniczne nr 1"],
        "class_tokens": ["Technik mechatronik", "ang-hisz"],
        "district": "Mokotów",
        "commute": "słaby",
        "notes": "Pierwszy wybór. Bardzo dobry kierunek pod politechnikę, ale dojazd z Kobyłki na Mokotów jest długi.",
    },
    {
        "source_order": 2,
        "school": "Technikum Mechatroniczne nr 1",
        "type": "Technikum",
        "class": "Technik mechatronik (ang-niem)",
        "school_tokens": ["Technikum Mechatroniczne nr 1"],
        "class_tokens": ["Technik mechatronik", "ang-niem"],
        "district": "Mokotów",
        "commute": "słaby",
        "notes": "Ten sam profil co pierwszy wybór; próg 2025 minimalnie niższy.",
    },
    {
        "source_order": 3,
        "school": "Technikum Mechatroniczne nr 1",
        "type": "Technikum",
        "class": "Technik programista (ang-niem)",
        "school_tokens": ["Technikum Mechatroniczne nr 1"],
        "class_tokens": ["Technik programista", "ang-niem"],
        "district": "Mokotów",
        "commute": "słaby",
        "notes": "Dobry pod informatykę/politechnikę, ale próg wyżej niż 168 w 2025.",
    },
    {
        "source_order": 4,
        "school": "Technikum Kinematograficzno-Komputerowe im. K. Kieślowskiego",
        "type": "Technikum",
        "class": "Technik programista (ang-niem)",
        "school_tokens": ["Kieślowskiego|Technikum Kinematograficzno-Komputerowe"],
        "class_tokens": ["Technik programista"],
        "district": "Śródmieście",
        "commute": "średni",
        "notes": "Bardzo sensowny techniczny wybór zapasowy przy profilu informatycznym.",
    },
    {
        "source_order": 5,
        "school": "VIII LO im. Władysława IV",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Władysława IV"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Świetne pod medycynę, ale bardzo ambitne przy 168 pkt.",
    },
    {
        "source_order": 6,
        "school": "CLVII LO im. Marii Skłodowskiej-Curie",
        "type": "LO",
        "class": "fiz-inf-mat",
        "school_tokens": ["CLVII", "Skłodowskiej"],
        "class_tokens": ["fiz-inf-mat"],
        "district": "Śródmieście",
        "commute": "średni",
        "notes": "Bardzo dobre dopasowanie pod politechnikę; próg w 2025 poniżej 168.",
    },
    {
        "source_order": 7,
        "school": "CLVII LO im. Marii Skłodowskiej-Curie",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["CLVII", "Skłodowskiej"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Śródmieście",
        "commute": "średni",
        "notes": "Również bardzo dobre pod politechnikę, trochę mniej informatyczne niż fiz-inf-mat.",
    },
    {
        "source_order": 8,
        "school": "VI LO im. Tadeusza Reytana",
        "type": "LO",
        "class": "fiz-ang-mat (ang-hisz)",
        "school_tokens": ["Reytana"],
        "class_tokens": ["fiz-ang-mat", "ang-hisz"],
        "district": "Mokotów",
        "commute": "słaby",
        "notes": "Merytorycznie mocne, ale daleko i na granicy/ambitnie zależnie od rocznika.",
    },
    {
        "source_order": 9,
        "school": "XIII LO im. płk. Leopolda Lisa-Kuli",
        "type": "LO",
        "class": "biol-chem-ang",
        "school_tokens": ["Leopolda Lisa-Kuli|Lisa-Kuli"],
        "class_tokens": ["biol-chem-ang"],
        "district": "Targówek",
        "commute": "dobry",
        "notes": "Bardzo dobry kompromis: blisko, medyczny kierunek, próg realny.",
    },
    {
        "source_order": 10,
        "school": "XII LO im. Henryka Sienkiewicza",
        "type": "LO",
        "class": "geogr-ang-mat",
        "school_tokens": ["Sienkiewicza"],
        "class_tokens": ["geogr-ang-mat"],
        "district": "Wola",
        "commute": "średni",
        "notes": "Profil bardziej ekonomiczno-techniczny niż medyczny; dobry zapas przy 168.",
    },
    {
        "source_order": 11,
        "school": "XIII LO im. płk. Leopolda Lisa-Kuli",
        "type": "LO",
        "class": "inf-ang-mat",
        "school_tokens": ["Leopolda Lisa-Kuli|Lisa-Kuli"],
        "class_tokens": ["inf-ang-mat"],
        "district": "Targówek",
        "commute": "dobry",
        "notes": "Jedna z najlepszych warszawskich opcji blisko Kobyłki pod informatykę/politechnikę.",
    },
    {
        "source_order": 12,
        "school": "XIII LO im. płk. Leopolda Lisa-Kuli",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Leopolda Lisa-Kuli|Lisa-Kuli"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Targówek",
        "commute": "dobry",
        "notes": "Blisko i dobrze dopasowane pod politechnikę.",
    },
    {
        "source_order": 13,
        "school": "VIII LO im. Władysława IV",
        "type": "LO",
        "class": "fiz-inf-mat",
        "school_tokens": ["Władysława IV"],
        "class_tokens": ["fiz-inf-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Bardzo dobre pod politechnikę, ale raczej ambitne przy 168.",
    },
    {
        "source_order": 14,
        "school": "VIII LO im. Władysława IV",
        "type": "LO",
        "class": "chem-fiz-mat",
        "school_tokens": ["Władysława IV"],
        "class_tokens": ["chem-fiz-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Najlepszy pomost politechnika/medycyna, próg minimalnie powyżej 168.",
    },
    {
        "source_order": 15,
        "school": "XIX LO im. Powstańców Warszawy",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Powstańców Warszawy"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Praga-Południe",
        "commute": "średni",
        "notes": "Dobre pod politechnikę, zwykle blisko granicy 168.",
    },
    {
        "source_order": 16,
        "school": "XXXV LO im. Bolesława Prusa",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Prusa"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Praga-Południe",
        "commute": "średni",
        "notes": "Dobra szkoła i profil, realność mocno zależy od rocznika.",
    },
    {
        "source_order": 17,
        "school": "CV LO im. Zbigniewa Herberta",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["CV", "Herberta"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Białołęka",
        "commute": "średni",
        "notes": "Bezpieczniejszy profil politechniczny, do rozważenia jako zapas.",
    },
    {
        "source_order": 18,
        "school": "CV LO im. Zbigniewa Herberta",
        "type": "LO",
        "class": "biol-chem-ang",
        "school_tokens": ["CV", "Herberta"],
        "class_tokens": ["biol-chem-ang"],
        "district": "Białołęka",
        "commute": "średni",
        "notes": "Bezpieczniejszy wariant medyczny niż Władysław IV.",
    },
    {
        "source_order": 19,
        "school": "Technikum Elektroniczne nr 1",
        "type": "Technikum",
        "class": "Technik mechatronik",
        "school_tokens": ["Technikum Elektroniczne nr 1"],
        "class_tokens": ["Technik mechatronik"],
        "district": "Wola",
        "commute": "średni",
        "notes": "Dobry techniczny zapas dla mechatroniki, próg niższy niż Wiśniowa.",
    },
    {
        "source_order": 20,
        "school": "Technikum Łączności",
        "type": "Technikum",
        "class": "Technik programista",
        "school_tokens": ["Technikum Łączności"],
        "class_tokens": ["Technik programista"],
        "district": "Praga-Południe",
        "commute": "średni",
        "notes": "Bezpieczny techniczny wybór, ale progi dużo niższe niż najlepsze technika.",
    },
    {
        "source_order": 21,
        "school": "XXVII LO im. Tadeusza Czackiego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Czackiego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardzo mocna śródmiejska opcja medyczna; wysoka selektywność i dobry dojazd metrem.",
    },
    {
        "source_order": 22,
        "school": "XXVII LO im. Tadeusza Czackiego",
        "type": "LO",
        "class": "fiz-inf-mat",
        "school_tokens": ["Czackiego"],
        "class_tokens": ["fiz-inf-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardzo mocna opcja pod politechnikę i informatykę; wysoki próg oznacza duże ryzyko.",
    },
    {
        "source_order": 23,
        "school": "IX LO im. Klementyny Hoffmanowej",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Hoffmanowej"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardzo mocny kierunek medyczny; sensowny jako ambitny wybór wysoko na liście.",
    },
    {
        "source_order": 24,
        "school": "IX LO im. Klementyny Hoffmanowej",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Hoffmanowej"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Mocny profil politechniczny z angielskim; bardzo ambitny punktowo.",
    },
    {
        "source_order": 25,
        "school": "IX LO im. Klementyny Hoffmanowej",
        "type": "LO",
        "class": "chem-fiz-mat",
        "school_tokens": ["Hoffmanowej"],
        "class_tokens": ["chem-fiz-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Dobry pomost między politechniką i kierunkami medyczno-przyrodniczymi.",
    },
    {
        "source_order": 26,
        "school": "II LO im. Stefana Batorego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Batorego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardzo mocna opcja medyczna w centrum; wysoki próg, ale dobry wybór preferencyjny.",
    },
    {
        "source_order": 27,
        "school": "II LO im. Stefana Batorego",
        "type": "LO",
        "class": "fiz-mat",
        "school_tokens": ["Batorego"],
        "class_tokens": ["fiz-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardzo mocna, dwujęzyczna opcja politechniczna; wymaga uwagi przy sprawdzianie językowym.",
    },
    {
        "source_order": 28,
        "school": "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika",
        "type": "LO",
        "class": "fiz-mat",
        "school_tokens": ["Kopernika"],
        "class_tokens": ["fiz-mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Bardzo mocny publiczny Kopernik pod politechnikę; klasa dwujęzyczna z wysokim progiem, więc warto dać wysoko jako ambitny wybór.",
    },
    {
        "source_order": 29,
        "school": "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Kopernika"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Jedna z najmocniejszych opcji medycznych w Warszawie; bardzo ambitna punktowo, ale zgodna z kierunkiem medycznym.",
    },
    {
        "source_order": 30,
        "school": "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika",
        "type": "LO",
        "class": "geogr-mat",
        "school_tokens": ["Kopernika"],
        "class_tokens": ["geogr", "mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Mocna matematyczna alternatywa w Koperniku; słabsze dopasowanie do medycyny, ale dobra jako ambitna klasa ogólnorozwojowa z matematyką.",
    },
    {
        "source_order": 31,
        "school": "XVIII LO im. Jana Zamoyskiego",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Zamoyskiego"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Mocna śródmiejska opcja politechniczna z dobrym dojazdem.",
    },
    {
        "source_order": 32,
        "school": "XVIII LO im. Jana Zamoyskiego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Zamoyskiego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Dobry wybór medyczny, trochę mniej ekstremalny niż Czacki/Hoffmanowa.",
    },
    {
        "source_order": 33,
        "school": "V LO im. Księcia Józefa Poniatowskiego",
        "type": "LO",
        "class": "chem-fiz-mat",
        "school_tokens": ["Poniatowskiego"],
        "class_tokens": ["chem-fiz-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Dobry profil pomostowy pod politechnikę i kierunki przyrodnicze.",
    },
    {
        "source_order": 34,
        "school": "V LO im. Księcia Józefa Poniatowskiego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Poniatowskiego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Mocny śródmiejski profil medyczny z dobrym dojazdem.",
    },
    {
        "source_order": 35,
        "school": "XXXVII LO im. Jarosława Dąbrowskiego",
        "type": "LO",
        "class": "fiz-inf-mat",
        "school_tokens": ["Dąbrowskiego"],
        "class_tokens": ["fiz-inf-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Sensowna opcja politechniczno-informatyczna przy bardzo dobrym dojeździe metrem.",
    },
    {
        "source_order": 36,
        "school": "XXXVII LO im. Jarosława Dąbrowskiego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Dąbrowskiego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Sensowny profil medyczny w centrum, mniej ryzykowny niż najbardziej selektywne licea.",
    },
    {
        "source_order": 37,
        "school": "XI LO im. Mikołaja Reja",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Reja"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Dobry profil politechniczny z angielskim i dojazdem przez centrum.",
    },
    {
        "source_order": 38,
        "school": "XI LO im. Mikołaja Reja",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Reja"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Dobry profil medyczny w Śródmieściu, warto rozważyć jako ambitny-realny wybór.",
    },
    {
        "source_order": 39,
        "school": "LXXXIII LO im. Emiliana Konopczyńskiego",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Konopczyńskiego"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardziej realna śródmiejska opcja politechniczna; dobry wariant zapasowy w centrum.",
    },
    {
        "source_order": 40,
        "school": "LXXXIII LO im. Emiliana Konopczyńskiego",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Konopczyńskiego"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Śródmieście",
        "commute": "dobry",
        "notes": "Bardziej realna śródmiejska opcja medyczna; dobra jako bezpieczniejszy wybór.",
    },
    {
        "source_order": 41,
        "school": "II LO im. Stefana Batorego - IB",
        "type": "LO IB",
        "class": "IB DP / pre-IB",
        "school_tokens": ["Batorego"],
        "class_tokens": ["[M]"],
        "district": "Śródmieście",
        "commute": "dobry",
        "ib": True,
        "ib_source": "https://batory.edu.pl/",
        "notes": "Publiczna ścieżka IB w bardzo mocnym liceum; sensowna tylko przy gotowości na intensywny angielski i szeroki program IB.",
    },
    {
        "source_order": 42,
        "school": "XXXIII LO Dwujęzyczne im. Mikołaja Kopernika - IB/MYP",
        "type": "LO IB",
        "class": "IB MYP / IB DP",
        "school_tokens": ["Kopernika"],
        "class_tokens": ["[M]"],
        "district": "Wola",
        "commute": "dobry",
        "ib": True,
        "ib_source": "https://kopernik.edu.pl/program-dyplomowy-matury-miedzynarodowej/",
        "notes": "Jedna z najmocniejszych publicznych opcji IB w Warszawie, bardzo dobra pod studia zagraniczne i ambitne STEM/med, ale ekstremalnie konkurencyjna.",
    },
    {
        "source_order": 43,
        "school": "XXXV LO im. Bolesława Prusa - IB",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Prusa"],
        "class_tokens": ["[M]"],
        "district": "Praga-Południe",
        "commute": "średni",
        "ib": True,
        "ib_source": "https://prus.edu.pl/ib-dp/program-ib-dp/",
        "notes": "Publiczna/ogólnodostępna ścieżka IB bliżej prawej strony Warszawy; dobra alternatywa, jeśli syn chce międzynarodowej matury bez rezygnacji z dojazdu.",
    },
    {
        "source_order": 44,
        "school": "2 SLO z Oddz. Międzynarodowymi im. P. Jasienicy STO",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Jasienicy"],
        "class_tokens": [],
        "district": "Śródmieście",
        "commute": "dobry",
        "ib": True,
        "ib_source": "https://ib.2slo.pl/",
        "notes": "Bardzo mocna społeczna szkoła IB; świetna pod studia zagraniczne, ale wymaga sprawdzenia kosztów, trybu rekrutacji i realnego wyboru przedmiotów HL.",
    },
    {
        "source_order": 45,
        "school": "Prywatne LO im. Zofii i Jędrzeja Moraczewskich Monnet International School",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Monnet"],
        "class_tokens": [],
        "district": "Mokotów",
        "commute": "słaby",
        "ib": True,
        "ib_source": "https://www.maturamiedzynarodowa.pl/liceum/",
        "notes": "Pełna prywatna ścieżka IB; dobra, jeśli priorytetem jest międzynarodowe środowisko, ale dojazd i koszt wymagają osobnej decyzji.",
    },
    {
        "source_order": 46,
        "school": "International American School of Warsaw",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["International American School"],
        "class_tokens": [],
        "district": "Ursynów",
        "commute": "słaby",
        "ib": True,
        "ib_source": "https://ias.edu.pl/",
        "notes": "Anglojęzyczna szkoła z IB; bardziej opcja międzynarodowa niż klasyczna ścieżka warszawskiej rekrutacji.",
    },
    {
        "source_order": 47,
        "school": "Warsaw Montessori High School",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Warsaw Montessori"],
        "class_tokens": [],
        "district": "Śródmieście",
        "commute": "dobry",
        "ib": True,
        "ib_source": "https://highschool.wmf.edu.pl/ib-diploma-programme/",
        "notes": "Prywatna, kameralna ścieżka IB w centrum; do sprawdzenia pod kątem poziomu HL z matematyki i nauk przyrodniczych.",
    },
    {
        "source_order": 48,
        "school": "Thames British School Warsaw",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Thames British"],
        "class_tokens": [],
        "district": "Mokotów",
        "commute": "słaby",
        "ib": True,
        "ib_source": "https://thamesbritishschool.pl/learning/diploma-program/",
        "notes": "Prywatna szkoła międzynarodowa z szeroką ofertą przedmiotów IB, w tym Math, Physics, Chemistry, Biology i Computer Science.",
    },
    {
        "source_order": 49,
        "school": "The British School Warsaw",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["The British School"],
        "class_tokens": [],
        "district": "Mokotów",
        "commute": "słaby",
        "ib": True,
        "ib_source": "https://www.nordangliaeducation.com/our-schools/warsaw/ib",
        "notes": "Bardzo międzynarodowa, kosztowna ścieżka IB; sensowna głównie przy planie studiów zagranicznych.",
    },
    {
        "source_order": 50,
        "school": "Międzynarodowe LO TE Vizja Warszawa Centrum",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["TE Vizja"],
        "class_tokens": [],
        "district": "Wola",
        "commute": "dobry",
        "ib": True,
        "ib_source": "https://tevizja.pl/warszawa-centrum/program-ib/",
        "notes": "Prywatna ścieżka IB w centrum; dobra logistycznie, ale wymaga sprawdzenia kosztów i jakości przedmiotów HL.",
    },
    {
        "source_order": 51,
        "school": "Prywatne LO Sióstr Nazaretanek z Oddziałami Międzynarodowymi",
        "type": "LO IB",
        "class": "IB DP",
        "school_tokens": ["Nazaretanek"],
        "class_tokens": [],
        "district": "Wilanów",
        "commute": "słaby",
        "ib": True,
        "ib_source": "https://nazaretanki.edu.pl/program-matury-miedzynarodowej-ib-dp/",
        "notes": "Prywatna ścieżka IB z pre-IB i DP; merytorycznie ciekawa, ale dojazd z Kobyłki jest słaby.",
    },
    {
        "source_order": 52,
        "school": "CLXI LO im. Władysława Bartoszewskiego",
        "type": "LO",
        "class": "geogr-ang-mat",
        "school_tokens": ["Bartoszewskiego"],
        "class_tokens": ["geogr-ang-mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Bufor poniżej 150 pkt w dobrej lokalizacji; profil mat-geogr-ang może pasować do kierunków techniczno-ekonomicznych.",
    },
    {
        "source_order": 53,
        "school": "CLXI LO im. Władysława Bartoszewskiego",
        "type": "LO",
        "class": "biol-chem-ang",
        "school_tokens": ["Bartoszewskiego"],
        "class_tokens": ["biol-chem-ang"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Bezpieczny bufor medyczno-przyrodniczy poniżej 150 pkt; do sprawdzenia poziom biologii i chemii.",
    },
    {
        "source_order": 54,
        "school": "XLVI LO im. Stefana Czarnieckiego",
        "type": "LO",
        "class": "biol-chem-obcy",
        "school_tokens": ["Czarnieckiego"],
        "class_tokens": ["biol-chem-obcy"],
        "district": "Targówek",
        "commute": "dobry",
        "notes": "Dobry lokalizacyjnie bufor z Targówka; profil przyrodniczy, ale mniej matematyczny niż idealny biol-chem-mat.",
    },
    {
        "source_order": 55,
        "school": "XLVI LO im. Stefana Czarnieckiego",
        "type": "LO",
        "class": "geogr-ang-mat",
        "school_tokens": ["Czarnieckiego"],
        "class_tokens": ["geogr-ang-mat"],
        "district": "Targówek",
        "commute": "dobry",
        "notes": "Bezpieczny profil matematyczno-geograficzny blisko Kobyłki; dobry jako dolne zabezpieczenie listy.",
    },
    {
        "source_order": 56,
        "school": "L LO im. Ruy Barbosy",
        "type": "LO",
        "class": "fiz-inf-mat",
        "school_tokens": ["Ruy Barbosy|Barbosy"],
        "class_tokens": ["fiz-inf-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Najbardziej sensowny bufor politechniczny poniżej 150 pkt: fizyka, informatyka i matematyka, dobra lokalizacja.",
    },
    {
        "source_order": 57,
        "school": "L LO im. Ruy Barbosy",
        "type": "LO",
        "class": "biol-chem-mat",
        "school_tokens": ["Ruy Barbosy|Barbosy"],
        "class_tokens": ["biol-chem-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Bezpieczny bufor medyczny poniżej 150 pkt w dobrej lokalizacji po prawej stronie Wisły.",
    },
    {
        "source_order": 58,
        "school": "XLV LO im. Romualda Traugutta",
        "type": "LO",
        "class": "inf-ang-mat",
        "school_tokens": ["Traugutta"],
        "class_tokens": ["inf-ang-mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Bufor informatyczno-matematyczny poniżej 150 pkt; dobra lokalizacja, ale warto zweryfikować poziom rozszerzeń.",
    },
    {
        "source_order": 59,
        "school": "LXXXVI LO im. Batalionu Zośka",
        "type": "LO",
        "class": "fiz-ang-mat",
        "school_tokens": ["Zośka|Zośki"],
        "class_tokens": ["fiz-ang-mat"],
        "district": "Wola",
        "commute": "dobry",
        "notes": "Bezpieczny bufor politechniczny poniżej 150 pkt; profil zgodny z matematyką i fizyką.",
    },
    {
        "source_order": 60,
        "school": "Technikum Geologiczno-Geodezyjno-Drogowe",
        "type": "Technikum",
        "class": "Technik programista",
        "school_tokens": ["Geologiczno-Geodezyjno-Drogowe|Kluźniaka"],
        "class_tokens": ["Technik programista"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Techniczny bufor poniżej 150 pkt w bardzo dobrej lokalizacji; mniej prestiżowy, ale zgodny z informatyką/politechniką.",
    },
    {
        "source_order": 61,
        "school": "LXXVI LO im. Marszałka Józefa Piłsudskiego",
        "type": "LO",
        "class": "geogr-ang-mat",
        "school_tokens": ["Piłsudskiego"],
        "class_tokens": ["geogr-ang-mat"],
        "district": "Praga-Północ",
        "commute": "dobry",
        "notes": "Bardzo bezpieczny lokalizacyjny bufor; profil matematyczny raczej pod kierunki techniczno-ekonomiczne niż medycynę.",
    },
]


def build_rows(thresholds: list[ThresholdRow], maturas: list[MaturaRow]) -> list[dict]:
    output_rows: list[dict] = []
    for candidate in CANDIDATES:
        matura = match_matura(maturas, candidate["school_tokens"])
        matches = match_candidate(thresholds, candidate["school_tokens"], candidate["class_tokens"])
        relaxed_matches = match_candidate(thresholds, candidate["school_tokens"], candidate["class_tokens"][:1])
        by_year: dict[int, list[ThresholdRow]] = {}
        for match in matches:
            if match.threshold is not None:
                by_year.setdefault(match.year, []).append(match)
        for year in (2023, 2024, 2025):
            if year not in by_year:
                fallback_rows = [match for match in relaxed_matches if match.year == year and match.threshold is not None]
                if fallback_rows:
                    by_year[year] = fallback_rows
        values = {}
        sources = {}
        for year in (2023, 2024, 2025):
            year_rows = by_year.get(year, [])
            if year_rows:
                values[year] = min(row.threshold for row in year_rows if row.threshold is not None)
                source_rows = [row for row in year_rows if row.threshold == values[year]]
                sources[year] = "; ".join(f"{row.source_file}: {row.class_name}" for row in source_rows[:2])
            else:
                values[year] = None
                sources[year] = ""
        latest = values[2025]
        latest_source_rows = by_year.get(2025) or by_year.get(2024) or by_year.get(2023) or []
        source_profile = latest_source_rows[0].class_name if latest_source_rows else candidate["class"]
        delta = round(ASSUMED_POINTS - latest, 2) if latest is not None else None
        poly, med = fit_score(candidate["class"], candidate["type"])
        p_score, m_score, fit_score_final, q_score, d_score, s_score, e_score, total_score = final_score(
            poly,
            med,
            latest,
            candidate["commute"],
            [values[2023], values[2024], values[2025]],
            matura,
            candidate.get("ib", False),
        )
        row = {
            "Kolejność źródłowa": candidate["source_order"],
            "Rekomendowana kolejność": None,
            "Szkoła": candidate["school"],
            "Typ": candidate["type"],
            "Dzielnica": candidate["district"],
            "Klasa/kierunek": candidate["class"],
            "Oznaczenie oddziału": class_marker(source_profile) or "brak",
            "Przedmioty z profilu/kierunku": profile_subjects(source_profile),
            "Przedmioty oceniane/punktowane": scored_subjects(source_profile),
            "Egzamin dodatkowy z języka angielskiego": language_exam(source_profile),
            "Matura międzynarodowa IB": "TAK" if candidate.get("ib", False) else "nie",
            "Próg 2023": values[2023] if values[2023] is not None else "brak analogicznego profilu w pliku",
            "Próg 2024": values[2024] if values[2024] is not None else "brak analogicznego profilu w pliku",
            "Próg 2025": values[2025] if values[2025] is not None else "brak analogicznego profilu w pliku",
            "Średnia progów": round(mean([v for v in values.values() if v is not None]), 2) if any(v is not None for v in values.values()) else None,
            "Luka do 168 pkt": delta,
            "Szansa przy 168 pkt": chance_label(delta),
            "Dojazd z Kobyłki": candidate["commute"],
            "Dopasowanie politechnika": poly,
            "Dopasowanie medycyna": med,
            "P - politechnika 0-100": p_score,
            "M - medycyna 0-100": m_score,
            "Dopasowanie strategiczne 0-100": fit_score_final,
            "E - matura szkoły 0-100": e_score,
            "Ranking maturalny 2026": matura.rank_2026 if matura and matura.rank_2026 is not None else "brak w danych",
            "Wskaźnik maturalny 2026": matura.overall if matura and matura.overall is not None else "brak w rankingu",
            "Wskaźnik matura obowiązkowa": matura.required if matura and matura.required is not None else "brak w rankingu",
            "Wskaźnik matura dodatkowa": matura.additional if matura and matura.additional is not None else "brak w rankingu",
            "Q - siła/selektywność klasy 0-100": q_score,
            "R - realność 0-100": admission_score(latest),
            "D - dojazd 0-100": d_score,
            "S - stabilność 0-100": s_score,
            "Atrakcyjność 0-100": total_score,
            "Wzór atrakcyjności": "0,30*DopasowanieStrategiczne + 0,25*E + 0,20*Q + 0,15*D + 0,10*S",
            "Priorytet": None,
            "Komentarz": candidate["notes"],
            "Źródło progów": " | ".join(source for source in sources.values() if source),
            "Źródło matury": f"{matura.source_file}: linia {matura.source_line}" if matura else "brak szkoły w warszawskich pozycjach rankingu maturalnego 2026",
            "Źródło IB": candidate.get("ib_source", ""),
        }
        row["Priorytet"] = total_score
        output_rows.append(row)
    output_rows.sort(key=lambda item: (-item["Priorytet"], item["Kolejność źródłowa"]))
    for index, row in enumerate(output_rows, start=1):
        row["Rekomendowana kolejność"] = index
    return output_rows


def style_sheet(ws, table_range: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = table_range
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 10,
        "B": 12,
        "C": 34,
        "D": 12,
        "E": 15,
        "F": 28,
        "G": 12,
        "H": 36,
        "I": 52,
        "J": 42,
        "K": 11,
        "L": 11,
        "M": 11,
        "N": 13,
        "O": 13,
        "P": 18,
        "Q": 16,
        "R": 16,
        "S": 16,
        "T": 10,
        "U": 55,
        "V": 80,
        "W": 14,
        "X": 14,
        "Y": 14,
        "Z": 14,
        "AA": 14,
        "AB": 14,
        "AC": 36,
        "AD": 10,
        "AE": 55,
        "AF": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_status_colors(ws) -> None:
    fills = {
        "bezpieczna": "C6EFCE",
        "realna": "D9EAD3",
        "na granicy": "FFF2CC",
        "ambitna": "FCE4D6",
        "bardzo ambitna": "F4CCCC",
    }
    headers = {cell.value: cell.column for cell in ws[1]}
    chance_col = headers.get("Szansa przy 168 pkt")
    commute_col = headers.get("Dojazd z Kobyłki")
    if not chance_col or not commute_col:
        return
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, chance_col).value
        fill_color = fills.get(label)
        if fill_color:
            ws.cell(row, chance_col).fill = PatternFill("solid", fgColor=fill_color)
        commute = ws.cell(row, commute_col).value
        if commute == "dobry":
            ws.cell(row, commute_col).fill = PatternFill("solid", fgColor="D9EAD3")
        elif commute == "średni":
            ws.cell(row, commute_col).fill = PatternFill("solid", fgColor="FFF2CC")
        elif commute == "słaby":
            ws.cell(row, commute_col).fill = PatternFill("solid", fgColor="FCE4D6")


def add_table(ws, name: str) -> None:
    table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    style_sheet(ws, table_range)


def write_workbook(rows: list[dict], thresholds: list[ThresholdRow], maturas: list[MaturaRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Od pierwszego wyboru"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in sorted(rows, key=lambda item: item["Kolejność źródłowa"]):
        ws.append([row[header] for header in headers])
    add_table(ws, "TabelaOdPierwszegoWyboru")
    apply_status_colors(ws)

    ranking = wb.create_sheet("Rekomendacja")
    ranking.append(headers)
    for row in rows:
        ranking.append([row[header] for header in headers])
    add_table(ranking, "TabelaRekomendacji")
    apply_status_colors(ranking)

    summary = wb.create_sheet("Podsumowanie")
    summary_rows = [
        ["Założenie", "Wartość"],
        ["Zakres", "Tylko szkoły ponadpodstawowe w Warszawie"],
        ["Przyjęta punktacja ucznia", ASSUMED_POINTS],
        ["Źródła progów", "Lokalne pliki Markdown z lat 2023, 2024 i 2025 w katalogu 'wyniki lata przeszle'"],
        ["Źródła matur", "Lokalne pliki ranking-licea-matura-2026.md/pdf oraz ranking-technika-matura-2026.md/pdf. Do punktacji brane są tylko pozycje z miejscowością Warszawa."],
        ["Interpretacja", "Najważniejszy ranking preferencji opiera się na atrakcyjności klasy. Ryzyko dostania się jest pokazane osobno."],
        ["Kolejność", "Rekomendowana kolejność sortuje po atrakcyjności, czyli promuje mocne i ambitne klasy wyżej, a bezpieczne wybory zostawia niżej."],
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 110
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].font = Font(color="FFFFFF", bold=True)
    summary["B1"].font = Font(color="FFFFFF", bold=True)

    algorithm = wb.create_sheet("Algorytm oceny")
    algorithm_rows = [
        ["Element", "Opis"],
        ["Atrakcyjność końcowa", "Atrakcyjność = 0,30*DopasowanieStrategiczne + 0,25*E + 0,20*Q + 0,15*D + 0,10*S"],
        ["Dopasowanie strategiczne", "0,75*max(P, M) + 0,25*min(P, M). Dzięki temu klasa bardzo dobra pod medycynę albo politechnikę może być wysoko."],
        ["E - matura szkoły", "Jakość maturalna szkoły: 0,60*wskaźnik maturalny 2026 + 0,20*wskaźnik matur obowiązkowych + 0,20*wskaźnik matur dodatkowych. Przy braku szkoły w warszawskich pozycjach rankingu przyjęto 45 pkt, żeby brak danych nie dominował nad profilem."],
        ["P - politechnika", "Dopasowanie politechniczne z arkusza w skali 0-5 pomnożone przez 20. Pełne P dostają m.in. mechatronik, automatyk, elektronik, informatyk/programista, fiz-mat, fiz-inf-mat, fiz-ang-mat i chem-fiz-mat."],
        ["M - medycyna", "Pełne M dostają biol-chem-mat oraz biol-chem. Chem-fiz-mat jest mocnym pomostem. Mechatronik/automatyk oraz fiz-inf-mat dostają umiarkowany komponent za ścieżki okołomedyczne: inżynieria biomedyczna, aparatura medyczna, robotyka medyczna."],
        ["Q - siła/selektywność klasy", "100 gdy próg 2025 >=178; 90 gdy >=172; 80 gdy >=168; 65 gdy >=162; 50 gdy >=155; 35 gdy <155."],
        ["D - dojazd", "100 dla dobrego dojazdu, 65 dla średniego, 30 dla słabego."],
        ["S - stabilność", "100 gdy rozpiętość dostępnych progów <=3 pkt; 75 gdy <=6; 50 gdy <=10; 25 gdy >10; 40 przy danych tylko z jednego roku."],
        ["R - realność", "Pokazywana osobno, nie obniża atrakcyjności. 100 gdy próg 2025 <=158; 85 gdy <=165; 70 gdy <=168; 55 gdy <=171; 35 gdy <=176; 15 gdy >176."],
        ["Założenie punktowe", f"Ryzyko jest liczone względem zakładanych {ASSUMED_POINTS} pkt i progu 2025."],
        ["Interpretacja", "Do listy preferencji ambitne i atrakcyjne klasy powinny być wysoko. Wyniki maturalne podnoszą jakość szkoły, ale nie przykrywają profilu, dojazdu i siły konkretnej klasy."],
    ]
    for row in algorithm_rows:
        algorithm.append(row)
    algorithm.column_dimensions["A"].width = 26
    algorithm.column_dimensions["B"].width = 120
    for row in algorithm.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    algorithm["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    algorithm["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    algorithm["A1"].font = Font(color="FFFFFF", bold=True)
    algorithm["B1"].font = Font(color="FFFFFF", bold=True)

    raw = wb.create_sheet("Dane źródłowe")
    raw.append(["Rok", "Dzielnica", "Szkoła", "Symbol", "Klasa", "Próg", "Plik"])
    relevant_school_names = {norm(candidate["school_tokens"][0]) for candidate in CANDIDATES}
    for row in thresholds:
        if row.threshold is None:
            continue
        school_norm = norm(row.school)
        class_norm = norm(row.class_name)
        if any(token in school_norm for token in relevant_school_names) or any(
            token in class_norm for token in ["fiz", "inf", "mat", "chem", "biol", "programista", "mechatronik", "informatyk"]
        ):
            raw.append([row.year, row.district, row.school, row.symbol, row.class_name, row.threshold, row.source_file])
    add_table(raw, "TabelaDanychZrodlowych")
    raw.column_dimensions["C"].width = 52
    raw.column_dimensions["E"].width = 46

    matura_raw = wb.create_sheet("Dane matury 2026")
    matura_raw.append(["Typ", "Ranking 2026", "Szkoła", "Miejscowość", "Wskaźnik", "Obowiązkowe", "Dodatkowe", "Plik", "Linia"])
    for row in maturas:
        if norm(row.city) != "warszawa":
            continue
        matura_raw.append(
            [
                row.school_type,
                row.rank_2026,
                row.school,
                row.city,
                row.overall,
                row.required,
                row.additional,
                row.source_file,
                row.source_line,
            ]
        )
    add_table(matura_raw, "TabelaMatury2026")
    matura_raw.column_dimensions["C"].width = 58
    matura_raw.column_dimensions["H"].width = 42

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)


def main() -> None:
    thresholds = parse_markdown_tables()
    maturas = parse_matura_rankings()
    rows = build_rows(thresholds, maturas)
    write_workbook(rows, thresholds, maturas)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()

